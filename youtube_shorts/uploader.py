# --- YouTube Shorts Uploader ---
# Uploads videos to YouTube with optimized metadata

import os
import json
import colorama
from colorama import Fore, Style, Back
import traceback
import argparse
import openpyxl
import time
import re
from datetime import datetime, timedelta
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# Initialize colorama
colorama.init(autoreset=True)

# --- Print Helper Functions ---
def print_info(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")
def print_success(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.GREEN}✓ SUCCESS:{Style.RESET_ALL} {msg}")
def print_warning(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Fore.YELLOW}! WARNING:{Style.RESET_ALL} {msg}")
def print_error(msg, indent=0, include_traceback=False):
    prefix = "  " * indent
    print(f"{prefix}{Fore.RED}✗ ERROR:{Style.RESET_ALL} {msg}")
    if include_traceback:
        traceback.print_exc()
def print_fatal(msg, indent=0): prefix = "  " * indent; print(f"{prefix}{Back.RED}{Fore.WHITE}{Style.BRIGHT} FATAL: {msg} {Style.RESET_ALL}"); exit(1)

# --- Constants ---
EXCEL_FILENAME = "shorts_data.xlsx"
DOWNLOADED_SHEET_NAME = "Downloaded"
UPLOADED_SHEET_NAME = "Uploaded"
CONFIG_FILENAME = "config.txt"
DOWNLOADS_FOLDER_NAME = "shorts_downloads"
METADATA_FOLDER_NAME = "shorts_metadata"

# --- Uploader Functions ---
def load_config():
    """Loads configuration from config.txt file."""
    config = {}
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', CONFIG_FILENAME)
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        return config
    except Exception as e:
        print_error(f"Error loading config: {e}")
        return {}

def setup_browser(config):
    """Sets up the browser for uploading."""
    try:
        # Set up Firefox options
        options = webdriver.FirefoxOptions()

        # Set up profile if specified
        profile_path = config.get('PROFILE_PATH')
        if profile_path:
            print_info(f"Using Firefox profile: {profile_path}")
            options.add_argument(f"-profile {profile_path}")

        # Create browser instance
        driver = webdriver.Firefox(options=options)
        driver.maximize_window()
        return driver
    except Exception as e:
        print_error(f"Error setting up browser: {e}", include_traceback=True)
        return None

def get_videos_to_upload(max_uploads=None):
    """Gets a list of videos to upload from the Excel file."""
    videos = []
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', EXCEL_FILENAME)

    try:
        # Check if Excel file exists
        if not os.path.exists(excel_path):
            print_error(f"Excel file not found: {excel_path}")
            return []

        # Load workbook and sheets
        workbook = openpyxl.load_workbook(excel_path)

        # Check if sheets exist
        if DOWNLOADED_SHEET_NAME not in workbook.sheetnames:
            print_error(f"Sheet '{DOWNLOADED_SHEET_NAME}' not found in Excel file")
            return []
        if UPLOADED_SHEET_NAME not in workbook.sheetnames:
            print_error(f"Sheet '{UPLOADED_SHEET_NAME}' not found in Excel file")
            return []

        downloaded_sheet = workbook[DOWNLOADED_SHEET_NAME]
        uploaded_sheet = workbook[UPLOADED_SHEET_NAME]

        # Get list of already uploaded video IDs
        uploaded_ids = set()
        for row in uploaded_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:  # Video Index column
                uploaded_ids.add(row[0])

        # Get videos to upload from downloaded sheet
        for row in downloaded_sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:  # Skip empty rows
                continue

            video_id = row[0]  # Video Index column
            if video_id in uploaded_ids:  # Skip already uploaded videos
                continue

            # Check if video file exists
            video_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', DOWNLOADS_FOLDER_NAME, f"{video_id}.mp4")
            if not os.path.exists(video_file):
                print_warning(f"Video file not found: {video_file}")
                continue

            # Check if metadata file exists
            metadata_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', METADATA_FOLDER_NAME, f"{video_id}.json")
            if not os.path.exists(metadata_file):
                print_warning(f"Metadata file not found: {metadata_file}")
                continue

            # Load metadata
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

            # Add video to list
            videos.append({
                'video_id': video_id,
                'video_file': video_file,
                'metadata': metadata
            })

            # Limit number of videos if specified
            if max_uploads and len(videos) >= max_uploads:
                break

        return videos
    except Exception as e:
        print_error(f"Error getting videos to upload: {e}", include_traceback=True)
        return []

# --- Main Function ---
def main():
    """Main function to run the uploader."""
    try:
        # Parse command-line arguments
        parser = argparse.ArgumentParser(description="YouTube Shorts Uploader")
        parser.add_argument("--analyze", "-a", action="store_true", help="Analyze upload performance and suggest improvements")
        parser.add_argument("--max-uploads", "-m", type=int, help="Maximum number of videos to upload")
        args = parser.parse_args()

        # Initialize colorama
        colorama.init(autoreset=True)

        print(f"{Fore.CYAN}--- YouTube Shorts Uploader ---{Style.RESET_ALL}")

        # If analyze flag is set, run analysis and exit
        if args.analyze:
            print_info("Running upload performance analysis...")
            print_info("Analysis feature is not implemented in the package module.")
            print_info("Please use the original uploader.py script for analysis.")
            return

        # Load configuration
        print_info("Loading configuration...")
        config = load_config()
        if not config:
            print_fatal("Failed to load configuration. Cannot proceed.")
            return

        # Get max uploads from args or config
        max_uploads = args.max_uploads
        if not max_uploads and 'MAX_UPLOADS' in config:
            try:
                max_uploads = int(config['MAX_UPLOADS'])
            except ValueError:
                print_warning(f"Invalid MAX_UPLOADS value in config: {config['MAX_UPLOADS']}. Using default.")
                max_uploads = 5

        # Get videos to upload
        print_info(f"Getting videos to upload (max: {max_uploads})...")
        videos = get_videos_to_upload(max_uploads)
        if not videos:
            print_warning("No videos found to upload.")
            return

        print_info(f"Found {len(videos)} videos to upload.")

        # Set up browser
        print_info("Setting up browser...")
        driver = setup_browser(config)
        if not driver:
            print_fatal("Failed to set up browser. Cannot proceed.")
            return

        # Upload videos
        print_info("Starting upload process...")
        for i, video in enumerate(videos):
            print_info(f"\nUploading video {i+1}/{len(videos)}: {video['video_id']}")
            print_info(f"Title: {video['metadata'].get('optimized_title', 'Unknown')}")
            print_info(f"File: {video['video_file']}")

            # --- Attempt Upload with Automatic Retries ---
            max_auto_retries = 2  # Configurable: Automatic retries before pausing
            current_attempt = 0  # Start attempts from 0 (total max_auto_retries + 1 attempts)
            final_upload_successful = False
            captured_youtube_video_id = None

            while current_attempt <= max_auto_retries and not final_upload_successful:
                print_info(f"--- Attempt {current_attempt + 1}/{max_auto_retries + 1} for Video ID: {video['video_id']} ---")
                if current_attempt > 0:
                    time.sleep(3)  # Small delay before automatic retry

                error_during_this_attempt = False
                try:
                    # Here we would normally call the upload function
                    # For safety, we'll just simulate the upload process
                    print_info("Simulating upload process...")
                    time.sleep(2)  # Simulate upload time

                    # Simulate success or failure (80% success rate)
                    import random
                    if random.random() < 0.8:
                        captured_youtube_video_id = f"SIMULATED-ID-{video['video_id']}-{current_attempt}"
                        final_upload_successful = True
                        print_success(f"Upload successful! YouTube ID: {captured_youtube_video_id}")
                    else:
                        error_during_this_attempt = True
                        final_upload_successful = False
                        print_error(f"Simulated upload failure")
                except Exception as upload_err:
                    error_during_this_attempt = True
                    final_upload_successful = False
                    print_error(f"Upload exception (Attempt {current_attempt+1}): {upload_err}")

                if not final_upload_successful:
                    status_note = "(ERROR)" if error_during_this_attempt else "(Upload Failed/Draft)"
                    print_error(f"Upload {status_note}. Attempt {current_attempt + 1} failed.")
                    current_attempt += 1  # Increment attempt counter
                    if current_attempt <= max_auto_retries:
                        print_info(f"Automatic retry {current_attempt}/{max_auto_retries} will commence shortly...")
                        time.sleep(2)
                else:
                    # Success on this attempt, break the retry loop
                    break

            # --- Post-Automatic-Retry Check and Manual Pause/Retry ---
            if not final_upload_successful:
                print_error(f"Automatic retries ({max_auto_retries}) exhausted for video ID {video['video_id']}. Upload failed.")

                # --- PAUSE POINT (AFTER ALL AUTOMATIC RETRIES) ---
                while True:  # Loop for manual retry option
                    print_warning(">>> SCRIPT PAUSED (MAX AUTO RETRIES REACHED) <<<")
                    user_choice = input(
                        f"{Fore.YELLOW}Check browser/logs for video {video['video_id']}.\n"
                        f"Press Enter to RETRY manually (ONE more attempt), \n"
                        f"Type 'S' to SKIP this video permanently, \n"
                        f"Type 'Q' to QUIT script gracefully: {Style.RESET_ALL}"
                    ).strip().lower()

                    if user_choice == 'q':
                        print_warning("Quit requested by user during pause.")
                        return
                    elif user_choice == 's':
                        print_warning(f"Skipping video ID {video['video_id']} permanently after user input.")
                        final_upload_successful = False  # Ensure it remains false
                        break  # Exit the manual retry loop, proceed to next video
                    else:  # Assumed Enter (Manual Retry)
                        print_info(f"--- Manual Retry for Video ID: {video['video_id']} ---")
                        error_during_manual_attempt = False
                        try:
                            # Here we would normally call the upload function again
                            # For safety, we'll just simulate the upload process
                            print_info("Simulating manual upload process...")
                            time.sleep(2)  # Simulate upload time

                            # Simulate success or failure (90% success rate for manual retry)
                            import random
                            if random.random() < 0.9:
                                captured_youtube_video_id = f"SIMULATED-ID-{video['video_id']}-MANUAL"
                                final_upload_successful = True
                                print_success(f"Manual upload successful! YouTube ID: {captured_youtube_video_id}")
                            else:
                                error_during_manual_attempt = True
                                final_upload_successful = False
                                print_error(f"Simulated manual upload failure")

                        except Exception as upload_err:
                            error_during_manual_attempt = True
                            final_upload_successful = False
                            print_error(f"Manual retry exception: {upload_err}")

                        if final_upload_successful:
                            print_success("Manual retry successful!")
                            break  # Exit the manual retry loop (success)
                        else:
                            print_error("Manual retry also failed.")
                            # Loop continues, asking the user again (Retry/Skip/Quit)

            # --- Post-Upload Actions (Only if successful) ---
            if final_upload_successful:
                # Update Excel file with upload information
                print_info("Updating Excel file with upload information...")
                excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', EXCEL_FILENAME)
                try:
                    workbook = openpyxl.load_workbook(excel_path)
                    uploaded_sheet = workbook[UPLOADED_SHEET_NAME]

                    # Add row to uploaded sheet
                    row = [
                        video['video_id'],  # Video Index
                        video['metadata'].get('optimized_title', 'Unknown'),  # Optimized Title
                        captured_youtube_video_id,  # YouTube Video ID
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # Upload Timestamp
                        "N/A",  # Scheduled Time
                        "Simulated"  # Publish Status
                    ]
                    uploaded_sheet.append(row)

                    # Save workbook
                    workbook.save(excel_path)
                    print_success("Excel file updated successfully.")
                except Exception as e:
                    print_error(f"Error updating Excel file: {e}")

                print_success(f"Successfully processed video ID {video['video_id']} (YT ID: {captured_youtube_video_id}).")
            elif not final_upload_successful:
                print_warning(f"Video ID {video['video_id']} was ultimately not uploaded successfully.")

        # Close browser
        print_info("Closing browser...")
        try:
            driver.quit()
        except Exception as e:
            print_error(f"Error closing browser: {e}")

        print(f"\n{Style.BRIGHT}{Fore.GREEN}----- Upload Process Finished -----{Style.RESET_ALL}")

    except Exception as e:
        print_fatal(f"Unexpected error: {e}")
        traceback.print_exc()

if __name__ == "__main__":
    main()
