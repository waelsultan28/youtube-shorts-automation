# --- START OF FILE performance_tracker.py ---

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
YouTube Performance Tracker

This script fetches performance metrics for uploaded videos using the YouTube Data API.
It reads the YouTube Video IDs from the Excel file and updates the metrics,
focusing only on videos scheduled within the last 7 days.
"""

import os
import json
import time
import pickle
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ... (Google API imports and Colorama setup remain the same) ...
try:
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    GOOGLE_API_AVAILABLE = True
except ImportError:
    print("Warning: Google API libraries not found. Install with:")
    print("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
    GOOGLE_API_AVAILABLE = False

try:
    import colorama
    from colorama import Fore, Style, init
    init(autoreset=True)
    COLOR_ENABLED = True
    print(f"{Fore.GREEN}Colorama loaded successfully. Colored output enabled.{Style.RESET_ALL}")
except ImportError:
    print("Warning: 'colorama' not found. Install it for colored output (`pip install colorama`). Output will be monochrome.")
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor()
    COLOR_ENABLED = False


# ... (Configuration constants remain the same) ...
script_directory = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = os.path.join(script_directory, "shorts_data.xlsx")
UPLOADED_SHEET_NAME = "Uploaded"
CLIENT_SECRETS_FILE = os.path.join(script_directory, "client_secret.json")
TOKEN_FILE = os.path.join(script_directory, "token.json")
ERROR_LOG_FILE = os.path.join(script_directory, "performance_tracker_error_log.txt")
SCOPES = ["https://www.googleapis.com/auth/youtube.readonly"]

# ... (Logging helper functions remain the same) ...
def sanitize_message(message: str) -> str:
    import re
    patterns = [
        (r'AIza[0-9A-Za-z\-_]{35}', 'API_KEY_REDACTED'),
        (r'(["\'])?(api[_-]?k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th|cr[e]dential)["\']?\s*[:=]\s*["\']?([^"\',\s]{8,})["\']?', r'\1\2\3=REDACTED'),
        (r'(https?://[^\s]+[?&][^\s]*(?:k[e]y|t[o]ken|s[e]cret|p[a]ssword|a[u]th)=[^\s&"]+)', r'URL_WITH_SENSITIVE_PARAMS_REDACTED'),
        (r'([\w\-]+\.)(k[e]y|p[e]m|c[e]rt|p12|pfx|p[a]ssword|t[o]ken|s[e]cret)', r'\1REDACTED'),
    ]
    sanitized = message
    for pattern, replacement in patterns:
        sanitized = re.sub(pattern, replacement, sanitized, flags=re.IGNORECASE)
    return sanitized

def log_error_to_file(message: str, include_traceback: bool = False):
    import traceback
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sanitized_message = sanitize_message(message)
    full_message = f"[{timestamp}] {sanitized_message}\n"
    if include_traceback:
        try:
            exc_info = traceback.format_exc()
            if exc_info and exc_info.strip() != 'NoneType: None':
                sanitized_traceback = sanitize_message(exc_info)
                full_message += sanitized_traceback + "\n"
        except Exception: pass
    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f: f.write(full_message)
    except Exception as e: print(f"CRITICAL: Failed to write to error log file '{ERROR_LOG_FILE}': {e}")

def print_section_header(title: str): print(f"\n{Style.BRIGHT}{Fore.CYAN}--- {title} ---{Style.RESET_ALL}")
def print_info(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.DIM}{Fore.BLUE}i INFO:{Style.RESET_ALL} {msg}")
def print_success(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {Fore.GREEN}{msg}{Style.RESET_ALL}")
def print_warning(msg: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {Fore.YELLOW}{msg}{Style.RESET_ALL}")
def print_error(msg: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = False):
    prefix = "  " * indent
    print(f"{prefix}{Style.BRIGHT}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {Fore.RED}{msg}{Style.RESET_ALL}")
    if log_to_file: log_error_to_file(f"ERROR: {msg}", include_traceback=include_traceback)

# ... (get_authenticated_service function remains the same) ...
def get_authenticated_service():
    if not GOOGLE_API_AVAILABLE: print_error("Google API libraries not available."); return None
    creds = None
    if os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE, 'rb') as token: creds = pickle.load(token)
            print_success("Cached credentials loaded.")
        except Exception as e: print_warning(f"Failed to load cached credentials: {e}. Will re-authenticate."); creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try: creds.refresh(Request()); print_success("Credentials refreshed successfully.")
            except Exception as e: print_warning(f"Failed to refresh credentials: {e}. Will perform new auth."); creds = None
        else:
            print_info("No valid cached credentials. Starting new authentication flow.")
            if not os.path.exists(CLIENT_SECRETS_FILE): print_error(f"FATAL: Client secrets file not found: {CLIENT_SECRETS_FILE}"); return None
            try:
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS_FILE, SCOPES)
                creds = flow.run_local_server(port=0); print_success("Authentication flow completed.")
            except Exception as e: print_error(f"Auth flow error: {e}", include_traceback=True); return None
        if creds and creds.valid:
            try:
                with open(TOKEN_FILE, 'wb') as token: pickle.dump(creds, token)
                print_success(f"Credentials saved to: {TOKEN_FILE}")
            except Exception as e: print_warning(f"Failed to save credentials: {e}")
    if creds and creds.valid:
        try: service = build('youtube', 'v3', credentials=creds); print_success("YouTube Data API service built."); return service
        except Exception as e: print_error(f"API service build error: {e}", include_traceback=True); return None
    else: print_error("Authentication failed."); return None

# ... (get_video_stats function remains the same) ...
def get_video_stats(service, video_id: str):
    try:
        response = service.videos().list(part="statistics", id=video_id).execute()
        if response and response.get('items'):
            stats = response['items'][0]['statistics']
            return {'viewCount': int(stats.get('viewCount', 0)), 'likeCount': int(stats.get('likeCount', 0)), 'commentCount': int(stats.get('commentCount', 0)), 'favoriteCount': int(stats.get('favoriteCount', 0)),}
        else: print_warning(f"Video ID {video_id} not found or no items returned."); return None
    except HttpError as e: print_error(f"API error fetching stats for {video_id}: {e}"); return None
    except Exception as e: print_error(f"Unexpected error fetching stats for {video_id}: {e}"); return None

# ... (update_excel_with_stats function remains the same) ...
def update_excel_with_stats(excel_path: str, sheet_name: str, stats_data: Dict[str, Dict]):
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames: print_error(f"Sheet '{sheet_name}' not found."); return False
        sheet = wb[sheet_name]; header = [cell.value for cell in sheet[1]]
        print_info(f"Loaded sheet '{sheet_name}'. Header: {header}")
        try:
            id_col_idx = views_col_idx = likes_col_idx = comments_col_idx = last_updated_col_idx = None
            required_new_headers = ["Views (YT)", "Likes (YT)", "Comments (YT)", "Last Updated"]
            header_map = {str(cell.value).strip().lower(): cell.column for cell in sheet[1] if cell.value is not None}
            id_col_idx = header_map.get('youtube video id')
            if id_col_idx is None: print_error("'YouTube Video ID' column not found."); return False
            next_col = len(header) + 1
            for required_header in required_new_headers:
                if required_header.lower() not in header_map:
                    print_info(f"Adding missing header: '{required_header}'")
                    sheet.cell(row=1, column=next_col, value=required_header); header_map[required_header.lower()] = next_col; next_col += 1
            views_col_idx = header_map.get("views (yt)"); likes_col_idx = header_map.get("likes (yt)"); comments_col_idx = header_map.get("comments (yt)"); last_updated_col_idx = header_map.get("last updated")
            if not all([views_col_idx, likes_col_idx, comments_col_idx, last_updated_col_idx]): print_error("Could not find necessary stat columns."); return False
        except Exception as e: print_error(f"Error finding columns: {e}", include_traceback=True); return False
        updated_count = 0; now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for row_idx in range(2, sheet.max_row + 1):
            video_id_cell = sheet.cell(row=row_idx, column=id_col_idx)
            youtube_id = str(video_id_cell.value).strip() if video_id_cell.value else None
            if youtube_id and youtube_id != "N/A" and youtube_id in stats_data:
                stats = stats_data[youtube_id]
                try:
                    sheet.cell(row=row_idx, column=views_col_idx, value=stats.get('viewCount'))
                    sheet.cell(row=row_idx, column=likes_col_idx, value=stats.get('likeCount'))
                    sheet.cell(row=row_idx, column=comments_col_idx, value=stats.get('commentCount'))
                    sheet.cell(row=row_idx, column=last_updated_col_idx, value=now_str)
                    updated_count += 1
                    print_info(f"Updated stats for {youtube_id}: V={stats.get('viewCount')}, L={stats.get('likeCount')}, C={stats.get('commentCount')}", indent=1)
                except Exception as e: print_error(f"Error updating row {row_idx} for {youtube_id}: {e}"); continue
        if updated_count > 0:
            print_success(f"Updated stats for {updated_count} videos.")
            try: wb.save(excel_path); print_success(f"Excel file saved: {excel_path}"); return True
            except PermissionError: print_error(f"PermissionError saving '{excel_path}'. Is it open?"); return False
            except Exception as e: print_error(f"Error saving Excel: {e}"); return False
        else: print_info("No videos found in sheet requiring stat updates."); return False # This message might be misleading now, changed below
    except FileNotFoundError: print_error(f"Excel file not found: {excel_path}"); return False
    except Exception as e: print_error(f"Unexpected error updating Excel: {e}", include_traceback=True); return False


# --- Main Function (MODIFIED) ---
def main():
    print_section_header("Starting YouTube Performance Tracker")

    if not GOOGLE_API_AVAILABLE:
        print_error("Google API libraries not installed. Please install required packages:")
        print_info("pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return

    service = get_authenticated_service()
    if not service:
        print_error("Could not authenticate with YouTube Data API. Exiting.")
        return

    print_info(f"Loading Excel file: {EXCEL_FILE_PATH}")
    videos_to_fetch: List[str] = []
    try:
        wb = load_workbook(EXCEL_FILE_PATH, read_only=True, data_only=True)
        if UPLOADED_SHEET_NAME not in wb.sheetnames:
            print_error(f"Sheet '{UPLOADED_SHEET_NAME}' not found in '{EXCEL_FILE_PATH}'. Exiting.")
            wb.close()
            return

        sheet = wb[UPLOADED_SHEET_NAME]
        header = [cell.value for cell in sheet[1]]

        # --- Find column indices (Case-insensitive check) ---
        id_col_idx = None
        schedule_time_col_idx = None # <-- Changed from upload_ts_col_idx

        header_map = {str(h).lower().strip(): i for i, h in enumerate(header, 1) if h}

        id_col_idx = header_map.get('youtube video id')
        schedule_time_col_idx = header_map.get('schedule time') # <-- Look for "Schedule Time"

        if id_col_idx is None:
            print_error("'YouTube Video ID' column not found in header. Cannot fetch stats.")
            wb.close()
            return

        if schedule_time_col_idx is None:
            # **Decision Point:** If Schedule Time is critical for filtering, exit?
            # Or proceed without filtering? For now, warn and proceed without filtering.
            print_warning("'Schedule Time' column not found in header. Cannot filter by date, will check ALL videos.")
            # If you absolutely need the filter, uncomment the next two lines:
            # print_error("'Schedule Time' column is required for date filtering. Exiting.")
            # wb.close(); return

        # --- Calculate the date 7 days ago ---
        now = datetime.now()
        seven_days_ago = now - timedelta(days=7)
        print_info(f"Checking for videos scheduled on or after: {seven_days_ago.strftime('%Y-%m-%d')}")
        # --- End date calculation ---

        print_info("Scanning for recent videos to update stats...")
        skipped_old = 0

        for row_idx in range(2, sheet.max_row + 1):
            youtube_id = None
            reference_datetime = None # Use a generic name for the date being checked

            try:
                # Get YouTube ID
                video_id_cell = sheet.cell(row=row_idx, column=id_col_idx)
                youtube_id = str(video_id_cell.value).strip() if video_id_cell.value else None

                if not youtube_id or youtube_id == "N/A":
                    continue

                # --- Get and Parse "Schedule Time" (only if column exists) ---
                if schedule_time_col_idx:
                    schedule_time_cell = sheet.cell(row=row_idx, column=schedule_time_col_idx)
                    schedule_time_value = schedule_time_cell.value

                    if not schedule_time_value or str(schedule_time_value).strip().upper() == "N/A":
                        # If schedule time is missing or N/A, we cannot filter by date.
                        # Decide whether to skip or include these. Let's include them for now.
                        print_info(f"Row {row_idx} (ID: {youtube_id}) - Missing or N/A schedule time. Including for stats check.", indent=1)
                        reference_datetime = None # Mark as cannot filter
                    else:
                        try:
                            # Handle Excel's date/time types or strings
                            if isinstance(schedule_time_value, float):
                                reference_datetime = datetime.fromtimestamp(time.mktime(time.gmtime((schedule_time_value - 25569) * 86400.0)))
                            elif isinstance(schedule_time_value, datetime):
                                reference_datetime = schedule_time_value
                            else: # Assume string format
                                reference_datetime = datetime.strptime(str(schedule_time_value), "%Y-%m-%d %H:%M:%S")

                            # --- Apply the date filter using Schedule Time ---
                            if reference_datetime < seven_days_ago:
                                skipped_old += 1
                                continue # Skip if scheduled older than 7 days
                            # --- End date filter ---

                        except (ValueError, TypeError):
                            print_warning(f"Skipping row {row_idx} (ID: {youtube_id}) - Could not parse Schedule Time '{schedule_time_value}'. Expected format 'YYYY-MM-DD HH:MM:SS' or Excel date.", indent=1)
                            continue # Skip if timestamp is invalid format
                # --- End Schedule Time Parsing ---

                # If schedule time column doesn't exist OR if the video is recent (or couldn't be filtered)
                # -> Add to fetch list regardless of existing stats
                if youtube_id not in videos_to_fetch: # Avoid duplicates
                    videos_to_fetch.append(youtube_id)
                    # Optional logging:
                    # date_str = reference_datetime.strftime('%Y-%m-%d') if reference_datetime else 'Unknown/Not Filtered'
                    # print_info(f"Adding video ID to fetch: {youtube_id} (Scheduled: {date_str})", indent=1)

            except Exception as row_err:
                print_error(f"Error processing row {row_idx}: {row_err}", indent=1, include_traceback=True)
                continue

        wb.close()

        if schedule_time_col_idx and skipped_old > 0: print_info(f"Skipped {skipped_old} videos scheduled older than 7 days.")
        elif not schedule_time_col_idx: print_warning("Could not filter by date (missing 'Schedule Time' column).")

    except FileNotFoundError:
        print_error(f"Excel file not found at: {EXCEL_FILE_PATH}. Exiting.")
        return
    except Exception as e:
        print_error(f"Error reading Excel file for IDs: {e}", include_traceback=True)
        return

    if not videos_to_fetch:
        print_info("No videos found scheduled within the last 7 days requiring stat updates.")
        return

    print_info(f"Found {len(videos_to_fetch)} videos scheduled within the last 7 days to fetch/update stats for.")

    # --- Fetch stats in batches (Keep existing batch logic) ---
    batch_size = 50
    all_fetched_stats: Dict[str, Dict] = {}

    for i in range(0, len(videos_to_fetch), batch_size):
        batch_ids = videos_to_fetch[i:i + batch_size]
        print_info(f"Fetching stats for batch {i//batch_size + 1}: {len(batch_ids)} videos.")

        try:
            response = service.videos().list(
                part="statistics",
                id=",".join(batch_ids)
            ).execute()

            if response and response.get('items'):
                for item in response['items']:
                    video_id = item.get('id')
                    stats = item.get('statistics')
                    if video_id and stats:
                        all_fetched_stats[video_id] = {
                            'viewCount': int(stats.get('viewCount', 0)),
                            'likeCount': int(stats.get('likeCount', 0)),
                            'commentCount': int(stats.get('commentCount', 0)),
                            'favoriteCount': int(stats.get('favoriteCount', 0)),
                        }
                print_success(f"Successfully fetched stats for {len(response['items'])} videos in batch.")
            elif response and 'items' in response and not response['items']:
                print_warning(f"API returned empty items list for batch {i//batch_size + 1}. IDs might be invalid/private.")
            else:
                print_error(f"API call for batch {i//batch_size + 1} failed or returned unexpected format.")

        except HttpError as e:
            print_error(f"API error fetching batch {i//batch_size + 1}: {e}")
        except Exception as e:
            print_error(f"Unexpected error fetching batch {i//batch_size + 1}: {e}")

        time.sleep(1)

    if all_fetched_stats:
        print_info(f"Total stats fetched for {len(all_fetched_stats)} unique videos.")
        update_excel_with_stats(EXCEL_FILE_PATH, UPLOADED_SHEET_NAME, all_fetched_stats)
    else:
        print_warning("No stats were successfully fetched for the recent videos identified.")

if __name__ == "__main__":
    main()

# --- END OF FILE ---