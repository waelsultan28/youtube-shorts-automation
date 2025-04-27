# --- Start of the script ---
import os
import time
import json
import re
import random
import csv # Keep import for potential future use, though not directly used for scheduling now
from datetime import datetime, timedelta, time as dt_time # Added time import
from typing import Optional, Tuple, List, Dict

import traceback # For detailed error logging to file
import platform # For OS detection
import subprocess # For running FFmpeg
import signal # For sending signals (like SIGINT equivalent) on non-Windows
import sys # For command-line arguments

# Import Google's Generative AI library for self-improvement features
try:
    import google.generativeai as genai
except ImportError:
    print("Warning: Google Generative AI library not found. Self-improvement features will be disabled.")
    print("To enable, install with: pip install google-generativeai")
    genai = None

# --- Colorama Setup ---
try:
    import colorama
    from colorama import Fore, Style, init
    init(autoreset=True) # Automatically reset style after each print
    COLOR_ENABLED = True
    print(f"{Fore.GREEN}Colorama loaded successfully. Colored output enabled.{Style.RESET_ALL}")
except ImportError:
    print("Warning: 'colorama' not found. Install it for colored output (`pip install colorama`). Output will be monochrome.")
    # Define dummy color objects if colorama is not available
    class DummyColor:
        def __getattr__(self, name): return ""
    Fore = DummyColor(); Style = DummyColor() # Assign instances
    COLOR_ENABLED = False
# --- End Colorama Setup ---

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    TimeoutException,
    WebDriverException,
    JavascriptException, # Added for clarity
    NoSuchWindowException, # Added for clarity
    InvalidSessionIdException, # Added for clarity
    NoSuchElementException # Needed for old select_date_in_calendar
)
from webdriver_manager.firefox import GeckoDriverManager
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --- Import YouTube Limits Module ---
try:
    from youtube_limits import (
        validate_description,
        validate_tags,
        DEFAULT_YOUTUBE_DESCRIPTION_LIMIT, # Import defaults
        DEFAULT_YOUTUBE_TAG_LIMIT,
        DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT,
        DEFAULT_YOUTUBE_MAX_TAGS_COUNT
    )
except ImportError:
    print(f"{Fore.RED}{Style.BRIGHT}ERROR:{Style.RESET_ALL}{Fore.YELLOW} Could not import from youtube_limits.py. Using fallback validation.")
    print(f"{Fore.YELLOW}       Ensure youtube_limits.py is in the same directory and defines required functions/constants for accurate validation.")
    # Dummy functions if module is missing
    DEFAULT_YOUTUBE_DESCRIPTION_LIMIT = 4900 # Use a slightly safer default
    DEFAULT_YOUTUBE_TAG_LIMIT = 100
    DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT = 450 # Use a slightly safer default
    DEFAULT_YOUTUBE_MAX_TAGS_COUNT = 40
    def validate_description(desc: str, limit: int = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT) -> Tuple[str, List[str]]:
        warnings = []
        validated_desc = desc
        if len(desc) > limit:
            warnings.append(f"Description truncated from {len(desc)} to {limit} characters.")
            validated_desc = desc[:limit]
        return validated_desc, warnings
    def validate_tags(tags: List[str], tag_char_limit: int = DEFAULT_YOUTUBE_TAG_LIMIT, total_char_limit: int = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT, max_count_limit: int = DEFAULT_YOUTUBE_MAX_TAGS_COUNT) -> Tuple[List[str], List[str]]:
        warnings = []
        validated_tags = []
        current_total_chars = 0
        for tag in tags:
            if not isinstance(tag, str) or not tag.strip():
                warnings.append(f"Skipped invalid/empty tag: '{tag}'")
                continue
            cleaned_tag = tag.strip()
            if len(validated_tags) >= max_count_limit:
                warnings.append(f"Tag count limit ({max_count_limit}) reached. Skipped tag: '{cleaned_tag[:30]}...'")
                continue # Stop adding tags if count limit reached
            if len(cleaned_tag) > tag_char_limit:
                warnings.append(f"Tag truncated from {len(cleaned_tag)} to {tag_char_limit} chars: '{cleaned_tag[:30]}...'")
                cleaned_tag = cleaned_tag[:tag_char_limit]
            # Check total char limit *before* adding the tag + comma (or just tag if first)
            potential_len = current_total_chars + len(cleaned_tag) + (1 if validated_tags else 0) # Add 1 for comma if not first tag
            if potential_len > total_char_limit:
                 warnings.append(f"Total tag character limit ({total_char_limit}) reached. Skipped tag: '{cleaned_tag[:30]}...'")
                 continue # Stop adding tags if total char limit reached
            validated_tags.append(cleaned_tag)
            current_total_chars += len(cleaned_tag) + (1 if len(validated_tags) > 1 else 0)
        return validated_tags, warnings
# --- End YouTube Limits Import ---

# --- Global Path Definitions ---
script_directory = os.path.dirname(os.path.abspath(__file__))
METADATA_FOLDER = os.path.join(script_directory, "shorts_metadata")
UPLOAD_FOLDER = os.path.join(script_directory, "shorts_downloads")
EXCEL_FILE_PATH = os.path.join(script_directory, "shorts_data.xlsx")
CONFIG_FILE_PATH = os.path.join(script_directory, "config.txt")
ERROR_LOG_FILE = os.path.join(script_directory, "upload_error_log.txt")
DEBUG_RECORDING_FOLDER = os.path.join(script_directory, "debug_recordings")
PERFORMANCE_METRICS_FILE = os.path.join(script_directory, "performance_metrics.json") # For tracking performance metrics
UPLOADER_ANALYSIS_LOG = os.path.join(script_directory, "uploader_analysis_log.txt") # For AI-generated analysis
UPLOAD_CORRELATION_CACHE_FILENAME = "upload_correlation_cache.json" # For tracking correlation between video index, discovery keyword, and YouTube Video ID
UPLOAD_CORRELATION_CACHE_PATH = os.path.join(script_directory, UPLOAD_CORRELATION_CACHE_FILENAME)
# --- End Path Definitions ---

# --- Error Types and Analysis Constants ---
ERROR_TYPES = {
    "title_input": "Finding/setting title input",
    "description_input": "Finding/setting description input",
    "show_more": "Finding/clicking 'Show more' button",
    "tags_input": "Finding/setting tags input",
    "next_button": "Finding/clicking 'Next' button",
    "public_radio": "Finding/selecting 'Public' radio button",
    "schedule_radio": "Finding/selecting 'Schedule' radio button",
    "date_picker": "Finding/setting date in calendar",
    "time_input": "Finding/setting time input",
    "publish_button": "Finding/clicking 'Publish' button",
    "schedule_button": "Finding/clicking 'Schedule' button",
    "confirmation": "Confirming upload completion",
    "browser_session": "Browser session issues",
    "other": "Other/unclassified errors"
}
MAX_ERROR_SAMPLES = 50  # Maximum number of error samples to keep in the metrics file
MIN_ERRORS_FOR_ANALYSIS = 10  # Minimum number of errors needed to trigger analysis
MIN_ERROR_RATE_FOR_ANALYSIS = 0.15  # Minimum error rate (15%) to trigger analysis
# --- End Error Types and Analysis Constants ---

# --- Global Variable for Active Recording Process (for cleanup) ---
_current_recording_process: Optional[subprocess.Popen] = None
_current_recording_filename: Optional[str] = None
# --- End Global Variable ---

# --- Logging Helper Functions (Enhanced Colors) ---
def log_error_to_file(message: str, error_type: str = "other", step: str = "unknown", video_index: str = "UNKNOWN", xpath: str = "", include_traceback: bool = False):
    """Logs a detailed error message to the error log file (plain text) with additional context.

    Args:
        message: The main error message to log
        error_type: The type of error (from ERROR_TYPES keys)
        step: The step in the upload process where the error occurred
        video_index: The index of the video being uploaded
        xpath: The XPath that was being used (if applicable)
        include_traceback: Whether to include the full traceback
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    full_message = f"[{timestamp}] {message}\n"

    # Add context information if provided
    if error_type != "other" or step != "unknown" or xpath:
        full_message += f"CONTEXT: Type={error_type}, Step={step}, VideoIndex={video_index}, XPath={xpath}\n"

    if include_traceback:
        try:
            exc_info = traceback.format_exc()
            # Only include traceback if it's meaningful
            if exc_info and exc_info.strip() != 'NoneType: None':
                 full_message += exc_info + "\n"
        except Exception: pass # Ignore errors during traceback formatting
    try:
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f: f.write(full_message)

        # Update performance metrics
        update_error_metrics(error_type, step, video_index, message, xpath)
    except Exception as e:
        # Use direct print as colored print might fail if colorama failed
        print(f"CRITICAL: Failed to write to error log file '{ERROR_LOG_FILE}': {e}")

def load_performance_metrics():
    """Loads performance metrics from the JSON file."""
    default_metrics = {
        "total_uploads_attempted": 0,
        "total_uploads_successful": 0,
        "total_errors": 0,
        "error_counts": {error_type: 0 for error_type in ERROR_TYPES.keys()},
        "error_samples": [],
        "runs": [],
        "last_analysis_date": ""
    }

    try:
        if os.path.exists(PERFORMANCE_METRICS_FILE):
            with open(PERFORMANCE_METRICS_FILE, "r", encoding="utf-8") as f:
                metrics = json.load(f)
            # Update with any missing keys from default
            for key, value in default_metrics.items():
                if key not in metrics:
                    metrics[key] = value
            # Ensure error_counts has all error types
            for error_type in ERROR_TYPES.keys():
                if error_type not in metrics["error_counts"]:
                    metrics["error_counts"][error_type] = 0
            return metrics
        else:
            return default_metrics
    except Exception as e:
        print(f"{Fore.YELLOW}Error loading performance metrics: {e}. Using default values.")
        return default_metrics

def save_performance_metrics(metrics):
    """Saves performance metrics to the JSON file."""
    try:
        with open(PERFORMANCE_METRICS_FILE, "w", encoding="utf-8") as f:
            json.dump(metrics, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print(f"{Fore.RED}Error saving performance metrics: {e}")

def update_error_metrics(error_type, step, video_index, error_message, xpath=""):
    """Updates the error metrics in the performance metrics file.

    Args:
        error_type: The type of error (from ERROR_TYPES keys)
        step: The step in the upload process where the error occurred
        video_index: The index of the video being uploaded
        error_message: The main error message
        xpath: The XPath that was being used (if applicable)
    """
    try:
        metrics = load_performance_metrics()

        # Update error counts
        metrics["total_errors"] += 1
        if error_type in metrics["error_counts"]:
            metrics["error_counts"][error_type] += 1
        else:
            metrics["error_counts"][error_type] = 1

        # Add error sample
        error_sample = {
            "type": error_type,
            "step": step,
            "video_index": video_index,
            "message": error_message,
            "xpath": xpath,
            "timestamp": datetime.now().isoformat()
        }

        # Keep only the most recent MAX_ERROR_SAMPLES samples
        metrics["error_samples"].append(error_sample)
        if len(metrics["error_samples"]) > MAX_ERROR_SAMPLES:
            metrics["error_samples"] = metrics["error_samples"][-MAX_ERROR_SAMPLES:]

        save_performance_metrics(metrics)
    except Exception as e:
        print(f"{Fore.RED}Error updating error metrics: {e}")

def load_correlation_cache():
    """Loads the upload correlation cache from JSON file."""
    default_cache = []
    try:
        if os.path.exists(UPLOAD_CORRELATION_CACHE_PATH):
            with open(UPLOAD_CORRELATION_CACHE_PATH, "r", encoding="utf-8") as f:
                content = f.read()
                if not content:
                    return default_cache
                cache = json.loads(content)
                if not isinstance(cache, list):
                    print_warning(f"Correlation cache file '{UPLOAD_CORRELATION_CACHE_FILENAME}' has invalid format. Returning empty cache.")
                    return default_cache
                return cache
        else:
            print_info(f"Correlation cache file '{UPLOAD_CORRELATION_CACHE_FILENAME}' not found. Creating new cache.")
            return default_cache
    except json.JSONDecodeError:
        print_error(f"Error decoding JSON from correlation cache file '{UPLOAD_CORRELATION_CACHE_FILENAME}'. Returning empty cache.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading correlation cache: {e}")
        return default_cache

def save_correlation_cache(cache_data):
    """Saves the upload correlation cache to JSON file."""
    try:
        with open(UPLOAD_CORRELATION_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
        print_info(f"Saved correlation cache with {len(cache_data)} entries.")
    except Exception as e:
        print_error(f"Error saving correlation cache: {e}")

def add_to_correlation_cache(video_index_str, discovery_keyword, youtube_video_id):
    """Adds a new entry to the correlation cache.

    Args:
        video_index_str: The video index string (e.g., "video1")
        discovery_keyword: The keyword used to discover/download the video
        youtube_video_id: The YouTube video ID after upload
    """
    if not youtube_video_id or not discovery_keyword:
        print_warning(f"Missing required data for correlation cache: YT ID: {youtube_video_id}, Keyword: {discovery_keyword}")
        return

    try:
        # Load existing cache
        cache = load_correlation_cache()

        # Create new entry
        new_entry = {
            "video_index": video_index_str,
            "discovery_keyword": discovery_keyword,
            "youtube_video_id": youtube_video_id,
            "added_timestamp": datetime.now().isoformat()
        }

        # Add to cache
        cache.append(new_entry)

        # Save updated cache
        save_correlation_cache(cache)
        print_success(f"Added correlation data for {video_index_str} with keyword '{discovery_keyword}' and YT ID: {youtube_video_id}")
    except Exception as e:
        print_error(f"Error adding to correlation cache: {e}")

def analyze_upload_errors_with_gemini():
    """Analyzes upload errors using Gemini AI and generates suggestions for improvement.

    Returns:
        A string containing the analysis and suggestions, or None if analysis failed.
    """
    try:
        # Load performance metrics
        metrics = load_performance_metrics()

        # Check if we have enough errors to analyze
        if metrics["total_errors"] < MIN_ERRORS_FOR_ANALYSIS:
            print(f"{Fore.YELLOW}Not enough errors ({metrics['total_errors']}) to perform analysis. Need at least {MIN_ERRORS_FOR_ANALYSIS}.")
            return None

        # Calculate error rate
        total_attempts = metrics["total_uploads_attempted"]
        if total_attempts == 0:
            print(f"{Fore.YELLOW}No upload attempts recorded. Cannot calculate error rate.")
            return None

        error_rate = metrics["total_errors"] / total_attempts
        if error_rate < MIN_ERROR_RATE_FOR_ANALYSIS:
            print(f"{Fore.YELLOW}Error rate ({error_rate:.1%}) is below threshold ({MIN_ERROR_RATE_FOR_ANALYSIS:.1%}). Analysis not needed.")
            return None

        # Read the error log file
        error_log_content = ""
        if os.path.exists(ERROR_LOG_FILE):
            with open(ERROR_LOG_FILE, "r", encoding="utf-8") as f:
                error_log_content = f.read()

        # Generate a summary of the performance metrics
        summary = ["=== Performance Summary ==="]
        summary.append(f"Total upload attempts: {metrics['total_uploads_attempted']}")
        summary.append(f"Total successful uploads: {metrics['total_uploads_successful']}")
        summary.append(f"Total errors: {metrics['total_errors']}")
        summary.append(f"Overall success rate: {metrics['total_uploads_successful'] / max(1, metrics['total_uploads_attempted']):.1%}")

        # Add error type breakdown
        summary.append("\n=== Error Type Breakdown ===")
        for error_type, count in metrics["error_counts"].items():
            if count > 0:
                summary.append(f"{ERROR_TYPES.get(error_type, error_type)}: {count} ({count / max(1, metrics['total_errors']):.1%})")

        # Add recent error samples
        if metrics["error_samples"]:
            summary.append("\n=== Recent Error Samples ===")
            for i, sample in enumerate(metrics["error_samples"][-5:], 1):
                summary.append(f"Sample {i}:")
                summary.append(f"  Type: {sample['type']}")
                summary.append(f"  Step: {sample['step']}")
                summary.append(f"  Message: {sample['message']}")
                if sample.get('xpath'):
                    summary.append(f"  XPath: {sample['xpath']}")

        performance_summary = "\n".join(summary)

        # Create the prompt for Gemini
        prompt = f"""
        Analyze the following performance data and error logs from a YouTube Shorts uploader script that uses Selenium WebDriver to automate uploads.
        The script navigates through YouTube Studio's upload flow, sets metadata, and either publishes immediately or schedules videos.

        Performance Summary:
        {performance_summary}

        Recent Error Log Entries:
        {error_log_content[-5000:] if len(error_log_content) > 5000 else error_log_content}

        Based on this data, please provide:

        1. Analysis of the most common or critical failure points in the upload process.
        2. Potential reasons why these errors might be occurring (e.g., XPath selectors outdated, network issues, YouTube UI changes).
        3. Specific technical suggestions to improve reliability, such as:
           - Alternative XPath selectors to try
           - Timeout adjustments for specific steps
           - Additional retry logic for problematic steps
           - Handling of specific error conditions
        4. Configuration parameter suggestions (if applicable)

        Format your response clearly with headings and bullet points for easy reading.
        """

        # Call Gemini API
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        analysis = response.text.strip()

        # Save the analysis to the log file
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(UPLOADER_ANALYSIS_LOG, "a", encoding="utf-8") as f:
            f.write(f"\n\n=== Analysis Generated on {timestamp} ===\n\n")
            f.write(analysis)
            f.write("\n\n=== End of Analysis ===\n")

        # Update the last analysis date in the metrics
        metrics["last_analysis_date"] = timestamp
        save_performance_metrics(metrics)

        return analysis
    except Exception as e:
        print(f"{Fore.RED}Error analyzing upload errors: {e}")
        traceback.print_exc()
        return None

def print_section_header(title: str): print(f"\n{Style.BRIGHT}{Fore.CYAN}--- {title} ---{Style.RESET_ALL}")
def print_info(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.DIM}{Fore.BLUE}i INFO:{Style.RESET_ALL} {message}")
def print_success(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.GREEN}OK SUCCESS:{Style.RESET_ALL} {Fore.GREEN}{message}{Style.RESET_ALL}")
def print_warning(message: str, indent: int = 0): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.YELLOW}WARN WARNING:{Style.RESET_ALL} {Fore.YELLOW}{message}{Style.RESET_ALL}"); # Optionally log warnings to file: log_error_to_file(f"Warning: {message}")
def print_error(message: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = False): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.RED}ERR ERROR:{Style.RESET_ALL} {Fore.RED}{message}{Style.RESET_ALL}"); log_error_to_file(f"ERROR: {message}", include_traceback=include_traceback) if log_to_file else None
def print_fatal(message: str, indent: int = 0, log_to_file: bool = True, include_traceback: bool = True): prefix = "  " * indent; print(f"{prefix}{Style.BRIGHT}{Fore.RED}FATAL ERROR:{Style.RESET_ALL} {Fore.RED}{message}{Style.RESET_ALL}"); log_error_to_file(f"FATAL: {message}", include_traceback=include_traceback) if log_to_file else None
def print_config(key: str, value: any): print(f"  {Fore.MAGENTA}{key:<28}:{Style.RESET_ALL} {Style.BRIGHT}{value}{Style.RESET_ALL}")
# --- End Logging Helper Functions ---

# --- Configuration Loading ---
config = {}
try:
    print_info(f"Loading configuration from: {CONFIG_FILE_PATH}")
    with open(CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and "=" in line:
                key, value = line.split("=", 1)
                config[key.strip()] = value.strip()
    print_success("Configuration loaded.")
except FileNotFoundError: print_fatal(f"Configuration file '{CONFIG_FILE_PATH}' not found. Cannot continue.", log_to_file=False); raise
except Exception as e: print_fatal(f"Error reading configuration file '{CONFIG_FILE_PATH}': {e}. Cannot continue.", log_to_file=False); raise
# --- End Configuration Loading ---


# --- Get Configurable Settings ---
# General settings
_DEFAULT_MAX_UPLOADS = 25
_DEFAULT_CATEGORY = "Gaming"

try: max_uploads = int(config.get("MAX_UPLOADS", _DEFAULT_MAX_UPLOADS)); assert max_uploads > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid MAX_UPLOADS in config. Using default: {_DEFAULT_MAX_UPLOADS}"); max_uploads = _DEFAULT_MAX_UPLOADS

upload_category = config.get("UPLOAD_CATEGORY", _DEFAULT_CATEGORY).strip()

profile_path_config = config.get("PROFILE_PATH")

# --- Scheduling Mode Settings ---
_DEFAULT_SCHEDULING_MODE = "default_interval"
_DEFAULT_SCHEDULE_INTERVAL = 120
_DEFAULT_MIN_SCHEDULE_AHEAD = 20
_DEFAULT_CUSTOM_TIMES_STR = "9:00 AM, 3:00 PM" # Example default if setting is missing

scheduling_mode = config.get("SCHEDULING_MODE", _DEFAULT_SCHEDULING_MODE).strip().lower()
if scheduling_mode not in ["default_interval", "custom_tomorrow"]:
    print_warning(f"Invalid SCHEDULING_MODE '{scheduling_mode}'. Using default: '{_DEFAULT_SCHEDULING_MODE}'")
    scheduling_mode = _DEFAULT_SCHEDULING_MODE

try: schedule_interval_minutes = int(config.get("SCHEDULE_INTERVAL_MINUTES", _DEFAULT_SCHEDULE_INTERVAL)); assert schedule_interval_minutes > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid SCHEDULE_INTERVAL_MINUTES. Using default: {_DEFAULT_SCHEDULE_INTERVAL}"); schedule_interval_minutes = _DEFAULT_SCHEDULE_INTERVAL

# --- Parse CUSTOM_SCHEDULE_TIMES ---
custom_schedule_times_str = config.get("CUSTOM_SCHEDULE_TIMES", _DEFAULT_CUSTOM_TIMES_STR).strip()
parsed_config_times: List[dt_time] = [] # List to store parsed time objects
if scheduling_mode == 'custom_tomorrow' and custom_schedule_times_str: # Only parse if needed and present
    time_strings = [t.strip() for t in custom_schedule_times_str.split(',')]
    for time_str in time_strings:
        if not time_str: continue # Skip empty strings
        try:
            parsed_t = datetime.strptime(time_str, "%I:%M %p").time()
            parsed_config_times.append(parsed_t)
        except ValueError:
            try:
                 parsed_t = datetime.strptime(time_str, "%H:%M").time()
                 parsed_config_times.append(parsed_t)
                 print_warning(f"Parsed custom time '{time_str}' using 24-hour format (HH:MM). Recommended format is HH:MM AM/PM.", indent=1)
            except ValueError:
                 print_warning(f"Invalid format for custom schedule time '{time_str}' in config. Expected 'HH:MM AM/PM' or 'HH:MM'. Skipping this time.", indent=1)
    if parsed_config_times:
         print_success(f"Successfully parsed {len(parsed_config_times)} custom schedule times from config.", indent=1)
    else:
         print_warning("No valid custom schedule times found in config setting 'CUSTOM_SCHEDULE_TIMES'. Mode 'custom_tomorrow' will rely on interval fallback.", indent=1)
else:
     if scheduling_mode == 'custom_tomorrow': print_info("Config setting 'CUSTOM_SCHEDULE_TIMES' is empty or invalid. Mode 'custom_tomorrow' will rely on interval fallback.", indent=1)
# Ensure times are sorted for sequential use
parsed_config_times.sort()
# --- End Parse CUSTOM_SCHEDULE_TIMES ---

try:
    min_schedule_ahead_minutes = int(config.get("MIN_SCHEDULE_AHEAD_MINUTES", _DEFAULT_MIN_SCHEDULE_AHEAD))
    assert min_schedule_ahead_minutes >= 5 # Enforce a minimum reasonable value
except (ValueError, TypeError, AssertionError):
    print_warning(f"Invalid MIN_SCHEDULE_AHEAD_MINUTES. Using default: {_DEFAULT_MIN_SCHEDULE_AHEAD}")
    min_schedule_ahead_minutes = _DEFAULT_MIN_SCHEDULE_AHEAD
# --- End Scheduling Mode Settings ---

# --- Debug Recording Settings ---
enable_debug_recording = config.get("ENABLE_DEBUG_RECORDING", "False").strip().lower() == 'true'
ffmpeg_path_config = config.get("FFMPEG_PATH", "ffmpeg").strip() # Default to 'ffmpeg' expecting it in PATH
# --- End Debug Recording Settings ---

# --- Gemini API Configuration ---
gemini_api_key = config.get("GEMINI_API_KEY", "").strip()
if gemini_api_key:
    try:
        genai.configure(api_key=gemini_api_key)
        print_success("Gemini API configured successfully.")
    except Exception as e:
        print_warning(f"Failed to configure Gemini API: {e}. Self-improvement features will be disabled.")
else:
    print_warning("GEMINI_API_KEY not found in config. Self-improvement features will be disabled.")
    print_info("To enable, add GEMINI_API_KEY=your_api_key to config.txt")
# --- End Gemini API Configuration ---

# --- Read YouTube Limits from Config ---
try: cfg_desc_limit = int(config.get("YOUTUBE_DESCRIPTION_LIMIT", DEFAULT_YOUTUBE_DESCRIPTION_LIMIT)); assert cfg_desc_limit > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid YOUTUBE_DESCRIPTION_LIMIT in config. Using default: {DEFAULT_YOUTUBE_DESCRIPTION_LIMIT}"); cfg_desc_limit = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT

try: cfg_tag_limit = int(config.get("YOUTUBE_TAG_LIMIT", DEFAULT_YOUTUBE_TAG_LIMIT)); assert cfg_tag_limit > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid YOUTUBE_TAG_LIMIT in config. Using default: {DEFAULT_YOUTUBE_TAG_LIMIT}"); cfg_tag_limit = DEFAULT_YOUTUBE_TAG_LIMIT

try: cfg_total_tags_limit = int(config.get("YOUTUBE_TOTAL_TAGS_LIMIT", DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT)); assert cfg_total_tags_limit > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid YOUTUBE_TOTAL_TAGS_LIMIT in config. Using default: {DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT}"); cfg_total_tags_limit = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT

try: cfg_max_tags_count = int(config.get("YOUTUBE_MAX_TAGS_COUNT", DEFAULT_YOUTUBE_MAX_TAGS_COUNT)); assert cfg_max_tags_count > 0
except (ValueError, TypeError, AssertionError): print_warning(f"Invalid YOUTUBE_MAX_TAGS_COUNT in config. Using default: {DEFAULT_YOUTUBE_MAX_TAGS_COUNT}"); cfg_max_tags_count = DEFAULT_YOUTUBE_MAX_TAGS_COUNT
# --- End Reading YouTube Limits ---

# --- End Configurable Settings ---


def mimic_human_action_delay(min_sec=0.3, max_sec=0.8): time.sleep(random.uniform(min_sec, max_sec))

def configure_driver() -> Optional[webdriver.Firefox]:
    print_info("Configuring Firefox WebDriver...")
    firefox_options = FirefoxOptions()
    firefox_options.add_argument("--start-maximized")
    firefox_options.add_argument("--disable-notifications")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0" # Example, keep reasonably updated
    firefox_options.set_preference("general.useragent.override", user_agent)
    firefox_options.set_preference("media.peerconnection.enabled", False)
    firefox_options.set_preference("geo.enabled", False)
    firefox_options.add_argument('--disable-blink-features=AutomationControlled')

    if profile_path_config:
        profile_path = profile_path_config.strip()
        if os.path.isdir(profile_path):
            try:
                # Note: Using Options.profile is preferred over deprecated FirefoxProfile
                firefox_options.profile = profile_path # Directly assign the path string
                print_info(f"Using Firefox profile: {profile_path}", indent=1)
            except Exception as e: msg = f"Error setting specified Firefox profile path '{profile_path}'. Trying default profile. Error: {e}"; print_warning(msg, indent=1); log_error_to_file(f"Warning: {msg}")
        else: msg = f"Specified PROFILE_PATH '{profile_path}' not found. Using default profile."; print_warning(msg, indent=1); log_error_to_file(f"Warning: {msg}")
    else: print_info("Using default Firefox profile (PROFILE_PATH not set).", indent=1)

    driver = None
    try:
        print_info("Setting up GeckoDriver using webdriver-manager...", indent=1)
        geckodriver_log_path = os.path.join(script_directory, "geckodriver.log")
        try:
            service = FirefoxService(executable_path=GeckoDriverManager().install(), log_path=geckodriver_log_path)
        except Exception as e:
            print_warning(f"Could not set geckodriver log path '{geckodriver_log_path}': {e}. Using default log path.", indent=2)
            service = FirefoxService(executable_path=GeckoDriverManager().install())

        driver = webdriver.Firefox(service=service, options=firefox_options)
        print_success("WebDriver setup complete.", indent=1); print_info(f"GeckoDriver log: {geckodriver_log_path}", indent=2)
        return driver
    except WebDriverException as e:
        if "connection refused" in str(e).lower(): msg = f"WebDriver setup failed: Connection Refused. Is Firefox running or blocked by firewall? Error: {e}"
        elif "binary" in str(e).lower(): msg = f"WebDriver setup failed: Firefox Binary issue. Is Firefox installed correctly? Error: {e}"
        else: msg = f"WebDriver setup failed: {e}\n{Fore.YELLOW}       Check Firefox installation, network, driver/browser compatibility.{Style.RESET_ALL}"
        print_error(msg, indent=1, include_traceback=False); log_error_to_file(f"WebDriver setup failed: {e}", include_traceback=True); raise
    except Exception as e: msg = f"Unexpected error during WebDriver setup: {e}"; print_error(msg, indent=1, include_traceback=False); log_error_to_file(f"Unexpected error during WebDriver setup: {e}", include_traceback=True); raise


def update_excel_data(downloaded_sheet: Worksheet, uploaded_sheet: Worksheet, video_index_str: str, optimized_title: str, upload_time: datetime, schedule_time: Optional[datetime] = None, publish_status: str = "Published", youtube_video_id: Optional[str] = None):
    """Moves entry from Downloaded to Uploaded sheet in memory and adds YouTube Video ID if available."""
    print_info(f"Updating Excel data in memory for video index: {video_index_str}", indent=1)
    try:
        rows_to_delete = []; row_found_in_downloaded = False
        for row_idx in range(downloaded_sheet.max_row, 1, -1):
            cell = downloaded_sheet.cell(row=row_idx, column=1)
            if cell.value is not None and str(cell.value).strip().lower() == f"video{video_index_str}".lower(): rows_to_delete.append(row_idx); row_found_in_downloaded = True
        if row_found_in_downloaded:
            for row_idx in sorted(rows_to_delete, reverse=True): downloaded_sheet.delete_rows(row_idx)
            print_success(f"Removed entry for video{video_index_str} from 'Downloaded' sheet.", indent=2)
        else: print_warning(f"Could not find video{video_index_str} in 'Downloaded' sheet to remove.", indent=2)
        # Ensure consistent formatting for Excel
        upload_time_str = upload_time.strftime("%Y-%m-%d %H:%M:%S")
        schedule_time_str = schedule_time.strftime("%Y-%m-%d %H:%M:%S") if schedule_time else "N/A" # Use N/A if not scheduled
        # Include YouTube Video ID in the row data if available
        row_data = [ f"video{video_index_str}", optimized_title, youtube_video_id if youtube_video_id else "N/A", upload_time_str, schedule_time_str, publish_status ]
        uploaded_sheet.append(row_data)
        print_success(f"Appended entry to 'Uploaded' sheet (Status: {publish_status}, Schedule: {schedule_time_str}, YT ID: {youtube_video_id if youtube_video_id else 'N/A'}).", indent=2)
    except Exception as e: msg = f"Error updating Excel data in memory for video{video_index_str}: {e}"; print_error(msg, indent=1, include_traceback=True)


# --- Correlation Cache Functions ---
def load_correlation_cache():
    """Loads the upload correlation cache from JSON file."""
    default_cache = [] # List of dictionaries
    if not os.path.exists(UPLOAD_CORRELATION_CACHE_PATH):
        return default_cache
    try:
        with open(UPLOAD_CORRELATION_CACHE_PATH, "r", encoding="utf-8") as f:
            content = f.read()
            if not content: return default_cache # Handle empty file
            cache = json.loads(content)
            if not isinstance(cache, list):
                 print_warning(f"Correlation cache file '{UPLOAD_CORRELATION_CACHE_FILENAME}' has invalid format (expected list). Returning empty cache.")
                 return default_cache
            return cache
    except json.JSONDecodeError:
        print_error(f"Error decoding JSON from correlation cache file '{UPLOAD_CORRELATION_CACHE_FILENAME}'. Returning empty cache.")
        return default_cache
    except Exception as e:
        print_error(f"Error loading correlation cache: {e}", include_traceback=True)
        return default_cache

def save_correlation_cache(cache_data):
    """Saves the upload correlation cache to JSON file."""
    try:
        with open(UPLOAD_CORRELATION_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print_error(f"Error saving correlation cache: {e}", include_traceback=True)

def add_to_correlation_cache(video_index_str: str, discovery_keyword: Optional[str], youtube_video_id: str):
    """Adds a new entry to the correlation cache."""
    if not youtube_video_id: # Don't add if upload failed
        return
    cache = load_correlation_cache()
    new_entry = {
        "video_index": video_index_str,
        "discovery_keyword": discovery_keyword if discovery_keyword else "Unknown", # Handle missing keyword
        "youtube_video_id": youtube_video_id,
        "added_timestamp": datetime.now().isoformat() # Timestamp for cleanup
    }
    cache.append(new_entry)
    save_correlation_cache(cache)
    print_info(f"Added {video_index_str} (YT ID: {youtube_video_id}) to correlation cache.", indent=2)
# --- End Correlation Cache Functions ---


def check_and_update_scheduled(uploaded_sheet: Worksheet) -> bool:
    """Checks 'Uploaded' sheet and updates 'Scheduled' to 'Published' if time has passed."""
    print_info("Checking status of previously scheduled videos...")
    updated_count = 0; changes_made = False; now = datetime.now()
    if not uploaded_sheet or uploaded_sheet.max_row < 2: print_info("Uploaded sheet empty or only has headers. No check needed.", indent=1); return False

    # Corrected Column Indices: Schedule Time=5, Publish Status=6
    schedule_time_col_idx = 5
    publish_status_col_idx = 6
    video_id_col_idx = 1 # Keep as 1

    for row_idx in range(2, uploaded_sheet.max_row + 1):
        try:
            # Use the corrected column indices
            schedule_time_cell = uploaded_sheet.cell(row=row_idx, column=schedule_time_col_idx)
            publish_status_cell = uploaded_sheet.cell(row=row_idx, column=publish_status_col_idx)
            video_id_cell = uploaded_sheet.cell(row=row_idx, column=video_id_col_idx)

            video_id_str = str(video_id_cell.value) if video_id_cell.value else f"Row {row_idx}"

            # Check the correct status cell and that the schedule time cell has a value
            if publish_status_cell.value == "Scheduled" and schedule_time_cell.value and str(schedule_time_cell.value).strip().upper() != "N/A":
                schedule_time_str = str(schedule_time_cell.value)
                try:
                    # Handle potential floating point representation from Excel
                    if isinstance(schedule_time_cell.value, float):
                         schedule_time = datetime.fromtimestamp(time.mktime(time.gmtime((schedule_time_cell.value - 25569) * 86400.0)))
                    elif isinstance(schedule_time_cell.value, datetime):
                         schedule_time = schedule_time_cell.value
                    else: # Assume string format
                         schedule_time = datetime.strptime(schedule_time_str, "%Y-%m-%d %H:%M:%S")

                    # Compare current time with parsed schedule time
                    if now >= schedule_time:
                        publish_status_cell.value = "Published" # Update the correct status cell
                        print_success(f"Updated status for {video_id_str} to 'Published'.", indent=2)
                        updated_count += 1
                        changes_made = True
                except (ValueError, TypeError):
                     msg = f"Could not parse schedule time for {video_id_str}: Value='{schedule_time_str}', Format expected 'YYYY-MM-DD HH:MM:SS' or datetime object"; print_warning(msg, indent=2); log_error_to_file(f"Warning: {msg}"); continue
        except Exception as e:
             # Use video_id_col_idx here too for consistency
             video_id_cell = uploaded_sheet.cell(row=row_idx, column=video_id_col_idx); video_id_str = str(video_id_cell.value) if video_id_cell.value else f"Row {row_idx}"
             msg = f"Unexpected error checking schedule status for {video_id_str}: {e}"; print_error(msg, indent=2, include_traceback=False); log_error_to_file(msg, include_traceback=True); continue
    if updated_count > 0: print_info(f"Checked scheduled videos. Updated status for {updated_count} videos.", indent=1)
    else: print_info("Checked scheduled videos. No status updates needed based on current time.", indent=1)
    return changes_made


# --- REPLACED Function: select_date_in_calendar (from old script + FALLBACKS) ---
def select_date_in_calendar(driver, schedule_time):
    """Selects the date in the calendar using Backspace clearing + Fallback XPaths."""
    print_info("Selecting date in calendar (using old Backspace method + Fallbacks)...", indent=3)
    try:
        time.sleep(1)
        date_input_container_xpaths = [ "//tp-yt-paper-dialog[@id='dialog']//tp-yt-paper-input[@id='textbox']", ] # Add others if needed
        date_input_xpaths = [ ".//input", "//tp-yt-paper-dialog[@id='dialog']//input[contains(@class,'tp-yt-paper-input')]", ]
        date_input = None; container_found = False
        for container_xpath in date_input_container_xpaths:
             try:
                 print_info(f"Trying date container XPath: {container_xpath}", indent=4)
                 date_input_container = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, container_xpath)))
                 for input_xpath_rel in [".//input", ".//input[@id='input']"]:
                     try: date_input = date_input_container.find_element(By.XPATH, input_xpath_rel); print_success("Date input field located (within container).", indent=4); container_found = True; break
                     except NoSuchElementException: continue
                 if date_input: break
             except Exception: continue
        if not date_input:
             print_info("Could not find date input via container, trying direct XPaths...", indent=4)
             for xpath in date_input_xpaths:
                 try: print_info(f"Trying direct date input XPath: {xpath}", indent=4); date_input = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath))); print_success("Date input field located (direct XPath).", indent=4); break
                 except Exception: continue
        if not date_input: raise TimeoutException("Failed to find Date input field with any provided XPath.")
        date_string = schedule_time.strftime("%b %d, %Y"); print_info(f"Formatted date string (old format): '{date_string}'", indent=4)
        print_info("Clearing date input using Backspace x 15...", indent=4)
        for _ in range(15): date_input.send_keys(Keys.BACKSPACE); time.sleep(0.05)
        time.sleep(0.2); print_info("Backspace clearing complete.", indent=4)
        print_info(f"Sending date keys: '{date_string}'", indent=4); date_input.send_keys(date_string); time.sleep(0.5)
        print_info("Sending Enter key for date...", indent=4); date_input.send_keys(Keys.ENTER); time.sleep(1)
        print_success("Date selected using old method.", indent=4); return True
    except TimeoutException: msg = "Timeout finding date input container/input using old XPaths + Fallbacks."; print_error(msg, indent=3); log_error_to_file(f"ERROR: {msg}"); return False
    except Exception as e: msg = f"Error selecting date in calendar (using old method + Fallbacks): {e}"; print_error(msg, indent=3); log_error_to_file(f"ERROR: {msg}", include_traceback=True); return False
# --- END REPLACED Function ---


# --- Upload Video Function (Accepts Limits, with Fallbacks, User Schedule Radio XPaths) ---
def upload_video(
    driver: webdriver.Firefox,
    video_file: str,
    metadata: dict,
    publish_now: bool = True,
    schedule_time: Optional[datetime] = None,
    # --- Pass configured limits ---
    desc_limit: int = DEFAULT_YOUTUBE_DESCRIPTION_LIMIT,
    tag_char_limit: int = DEFAULT_YOUTUBE_TAG_LIMIT,
    total_char_limit: int = DEFAULT_YOUTUBE_TOTAL_TAGS_LIMIT,
    max_count_limit: int = DEFAULT_YOUTUBE_MAX_TAGS_COUNT
) -> Optional[str]:
    """Handles the video upload process on YouTube Studio, using provided limits and fallback XPaths."""
    video_index = metadata.get('video_index', 'UNKNOWN') # Get index for logging
    print_section_header(f"Uploading Video Index: {video_index}")
    print_info(f"File: {os.path.basename(video_file)}", indent=1)
    print_info(f"Title: {metadata.get('optimized_title', 'N/A')}", indent=1)

    upload_successful = False # Assume failure until explicitly set to True
    youtube_video_id: Optional[str] = None # Initialize YouTube Video ID

    # --- Validate Metadata ---
    print_info("Validating metadata...", indent=1)
    print_info(f"Limits Used: Desc={desc_limit}, TagChar={tag_char_limit}, TotalTags={total_char_limit}, MaxTags={max_count_limit}", indent=2)
    all_warnings = []
    if 'optimized_description' in metadata:
        desc = metadata['optimized_description']; validated_desc, desc_warnings = validate_description(desc, limit=desc_limit); metadata['optimized_description'] = validated_desc
        if desc_warnings: all_warnings.extend([f"Desc: {w}" for w in desc_warnings])
    else: metadata['optimized_description'] = ""; print_info("Metadata key 'optimized_description' missing.", indent=2)
    if 'optimized_tags' in metadata and isinstance(metadata['optimized_tags'], list):
        tags_list = metadata['optimized_tags']; validated_tags, tag_warnings = validate_tags(tags_list, tag_char_limit=tag_char_limit, total_char_limit=total_char_limit, max_count_limit=max_count_limit); metadata['optimized_tags'] = validated_tags
        if tag_warnings: all_warnings.extend([f"Tags: {w}" for w in tag_warnings])
    else: metadata['optimized_tags'] = []; print_info("Metadata key 'optimized_tags' missing or not list.", indent=2)
    if all_warnings: print_warning("Metadata validation finished with warnings:", indent=2); [print_warning(w, indent=3) for w in all_warnings]
    else: print_success("Metadata validated successfully.", indent=2)
    # --- End Validation ---

    try:
        # --- Navigation and File Selection ---
        print_info("Navigating to YouTube Studio...", indent=1)
        studio_url = "https://studio.youtube.com/"; driver.get(studio_url)
        wait_very_long = WebDriverWait(driver, 180); wait_long = WebDriverWait(driver, 60); wait_medium = WebDriverWait(driver, 45); wait_short = WebDriverWait(driver, 25)

        create_button_selector = "ytcp-button#create-icon, yt-icon-button#create-icon-button" # More robust selector
        create_button = wait_very_long.until(EC.element_to_be_clickable((By.CSS_SELECTOR, create_button_selector)), message="Timeout waiting for Create button")
        print_success("YouTube Studio page loaded, create button found.", indent=1)
        create_button.click(); mimic_human_action_delay(0.2, 0.5)

        upload_videos_selector = "tp-yt-paper-item#text-item-0"
        upload_button = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, upload_videos_selector)), message="Timeout waiting for 'Upload videos' option")
        upload_button.click(); mimic_human_action_delay(0.5, 1.0)

        file_input_xpath = "//input[@type='file']"
        file_input = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, file_input_xpath)), message="Timeout waiting for file input element")
        abs_video_path = os.path.abspath(video_file); print_info(f"Selecting file for upload: {abs_video_path}", indent=1)
        if not os.path.exists(abs_video_path): print_error(f"Video file does not exist at path: {abs_video_path}", indent=1); return False
        file_input.send_keys(abs_video_path)

        title_xpath_options = [
            "//ytcp-social-suggestions-textbox[@id='title-textarea']//div[@id='textbox']", # Preferred modern selector
            "//ytcp-mention-textbox[@label='Title (required)']//div[@id='textbox']",
            "//div[@id='textbox' and contains(@aria-label, 'Add a title that describes your video')]", # Fallback
        ]
        print_info("Waiting for upload details dialog to become active...", indent=1)
        try:
            title_elements_located = [EC.presence_of_element_located((By.XPATH, xp)) for xp in title_xpath_options]
            WebDriverWait(driver, 120).until(EC.any_of(*title_elements_located), message="Timeout waiting for upload dialog (any title element)")
        except TimeoutException as e: print_error(f"Timeout waiting for the upload details dialog (title input). Upload might have failed. {e}", indent=1); raise
        print_success("Upload details dialog active.", indent=1); mimic_human_action_delay(1.0, 1.5)
        # --- End Navigation ---

        # --- Fill Title ---
        print_info("Setting video title...", indent=2)
        title_input = None
        for i, xpath in enumerate(title_xpath_options):
            try:
                print_info(f"Trying title XPath {i+1}: {xpath}", indent=3)
                title_input = wait_short.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                print_success(f"Found title input with XPath {i+1}", indent=3); break
            except Exception:
                print_warning(f"Title XPath {i+1} failed, trying next...", indent=3)
                if i == len(title_xpath_options) - 1: raise TimeoutException("Failed to find title input with any XPath.")

        try:
            title_input.click(); mimic_human_action_delay(0.1, 0.3)
            title_input.send_keys(Keys.CONTROL + "a"); mimic_human_action_delay(0.05, 0.1)
            title_input.send_keys(Keys.DELETE); mimic_human_action_delay(0.1, 0.3)
            title_text = metadata.get("optimized_title", f"Video {video_index}");
            if not title_text: title_text = f"Video {video_index}" # Ensure fallback title if empty
            title_input.send_keys(title_text); mimic_human_action_delay(0.2, 0.4)
            # Send Space to attempt dismissing suggestions
            print_info("Sending Space to dismiss title suggestions...", indent=3)
            title_input.send_keys(Keys.SPACE); mimic_human_action_delay(0.2, 0.4)
            time.sleep(1.0) # Allow UI to settle
            # Verification
            try:
                 current_text = title_input.get_attribute('textContent').strip(); expected_text = title_text.strip()
                 if current_text != expected_text: print_warning(f"Title verification mismatch. Expected '{expected_text}', Got '{current_text}'.", indent=3); log_error_to_file(f"Warning: Title verification failed for video {video_index}. Expected '{expected_text}', got '{current_text}'")
                 else: print_success("Title content verified.", indent=3)
            except Exception as verif_err: print_warning(f"Could not verify title content: {verif_err}", indent=3)
            print_success("Title set.", indent=3)
        except TimeoutException: print_error("Timeout finding the Title input field.", indent=3); raise
        except Exception as e: print_error(f"Failed to set title: {e}", indent=3, include_traceback=False); log_error_to_file(f"Error setting title for video {video_index}: {e}", include_traceback=True); print_warning("Continuing despite title setting error.", indent=3)
        # --- End Title ---

        # --- Fill Description (JS method + Fallbacks + Added Delay) ---
        print_info("Setting video description (using JS method)...", indent=2)
        description_xpath_options = [
            "//ytcp-social-suggestions-textbox[@id='description-textarea']//div[@id='textbox']", # Preferred modern
            "//ytcp-mention-textbox[@label='Description']//div[@id='textbox']",
            "//div[@id='textbox' and contains(@aria-label, 'Tell viewers about your video')]", # Fallback
        ]
        description_input = None
        for i, xpath in enumerate(description_xpath_options):
            try:
                print_info(f"Trying description XPath {i+1}: {xpath}", indent=3)
                description_input = wait_short.until(EC.presence_of_element_located((By.XPATH, xpath)))
                print_success(f"Found description input with XPath {i+1}", indent=3); break
            except Exception:
                print_warning(f"Description XPath {i+1} failed, trying next...", indent=3)
                if i == len(description_xpath_options) - 1: raise TimeoutException("Failed to find description input with any XPath.")

        try:
            description_text = metadata.get("optimized_description", "")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", description_input); mimic_human_action_delay(0.2, 0.4)
            try: driver.execute_script("arguments[0].click();", description_input); mimic_human_action_delay(0.1, 0.3)
            except Exception as js_click_err: print_warning(f"JS click on description failed (might be okay): {js_click_err}", indent=3)

            print_info("Pausing briefly before setting description text...", indent=3); time.sleep(0.7) # Added delay

            driver.execute_script("arguments[0].textContent = arguments[1];", description_input, description_text); mimic_human_action_delay(0.3, 0.6)
            try: driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", description_input); mimic_human_action_delay(0.2, 0.4)
            except Exception as event_err: print_warning(f"Could not trigger input event for description: {event_err}", indent=3); log_error_to_file(f"Warning: Failed trigger input event description {video_index}: {event_err}")

            time.sleep(1.0) # Verification wait
            try:
                current_text = description_input.get_attribute('textContent'); expected_normalized = ' '.join(description_text.split()); actual_normalized = ' '.join(current_text.split())
                if actual_normalized != expected_normalized:
                    print_warning(f"Description verification mismatch (after delay). Check logs.", indent=3)
                    log_error_to_file(f"Warning: Description verification failed (JS Set+Delay) for video {video_index}. Expected approx: '{expected_normalized[:100]}...', Got approx: '{actual_normalized[:100]}...'")
                else: print_success("Description content verified (JS Set+Delay).", indent=3)
            except Exception as verif_err: print_warning(f"Could not verify description content: {verif_err}", indent=3)
            print_success("Attempted to set description using JavaScript (with delay).", indent=3)
        except TimeoutException: print_error("Timeout finding the Description input field.", indent=3); raise
        except Exception as e: print_error(f"Failed to set description: {e}", indent=3, include_traceback=False); log_error_to_file(f"Error setting description for video {video_index}: {e}", include_traceback=True); print_warning("Continuing despite description setting error.", indent=3)
        # --- End Description ---

        # --- Show More, Altered Content ---
        print_info("Handling 'Show More' and 'Altered Content'...", indent=2)
        show_more_xpath_options = [
            "//ytcp-button-shape//button[@aria-label='Show advanced settings']", # Preferred modern
            "//ytcp-button[@id='toggle-button']//div[contains(text(), 'Show more')]", # Alternative structure
            "//button[@aria-label='Show advanced settings']", # Simpler fallback
        ]
        show_more_clicked = False
        for i, xpath in enumerate(show_more_xpath_options):
             try:
                 print_info(f"Trying show more XPath {i+1}: {xpath}", indent=3); show_more_button = WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                 driver.execute_script("arguments[0].scrollIntoView({block:'center'});", show_more_button); time.sleep(0.3)
                 show_more_button.click(); mimic_human_action_delay(0.5, 1.0); print_success("'Show more' clicked.", indent=3); show_more_clicked = True; break
             except Exception: print_warning(f"Show more XPath {i+1} failed, trying next...", indent=3);
             if i == len(show_more_xpath_options) - 1: print_info("'Show more' button not found or already expanded.", indent=3)

        altered_content_no_xpath_options = [
            "//tp-yt-paper-radio-button[@name='VIDEO_HAS_ALTERED_CONTENT_NO']", # Preferred name attribute
            "//tp-yt-paper-radio-button[contains(@aria-label, 'No, it') and contains(@aria-label, 'altered content')]", # Aria label fallback
        ]
        altered_content_section_xpath = "//ytcp-video-metadata-altered-content"
        try:
             WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, altered_content_section_xpath)))
             print_info("Altered content section found.", indent=3); no_altered_content_label = None
             for i, xpath in enumerate(altered_content_no_xpath_options):
                 try:
                     print_info(f"Trying Altered Content 'No' XPath {i+1}: {xpath}", indent=3)
                     no_altered_content_label = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                     print_success(f"Found Altered Content 'No' with XPath {i+1}", indent=3); break
                 except Exception: print_warning(f"Altered Content 'No' XPath {i+1} failed, trying next...", indent=3);
                 if i == len(altered_content_no_xpath_options) - 1: print_info("Altered Content 'No' button not found.", indent=3); raise TimeoutException("No button found")
             driver.execute_script("arguments[0].scrollIntoView({block:'center'});", no_altered_content_label); time.sleep(0.3)
             if no_altered_content_label.get_attribute("aria-checked") != "true": no_altered_content_label.click(); mimic_human_action_delay(0.2, 0.4); print_success("'No' for altered content selected.", indent=3)
             else: print_info("'No' for altered content already selected.", indent=3)
        except TimeoutException: print_info("'Altered content' section/option not found or needed.", indent=3)
        except Exception as e: print_warning(f"Minor error interacting with 'Altered content': {e}", indent=3); log_error_to_file(f"Warning: Minor error altered content {video_index}: {e}")
        # --- End Show More / Altered ---

        # --- Add Tags (Word by Word) ---
        print_info("Adding tags...", indent=2)
        tags_input_xpath_options = [
            "//ytcp-form-input-container[@id='tags-container']//input[@id='text-input']", # Preferred modern
            "//input[contains(@aria-label, 'Tags') and @id='text-input']", # Aria label + ID
            "//ytcp-free-text-chip-bar//input[@id='text-input']", # Older structure fallback
        ]
        tags_input = None
        for i, xpath in enumerate(tags_input_xpath_options):
            try:
                print_info(f"Trying tags input XPath {i+1}: {xpath}", indent=3)
                tags_input = WebDriverWait(driver, 7).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                print_success(f"Found tags input with XPath {i+1}", indent=3); break
            except Exception:
                print_warning(f"Tags input XPath {i+1} failed, trying next...", indent=3)
                if i == len(tags_input_xpath_options) - 1: raise TimeoutException("Failed to find tags input with any XPath.")

        try:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", tags_input); time.sleep(0.5)
            validated_tags = metadata.get("optimized_tags", [])
            if validated_tags:
                tags_added_count = 0
                for tag in validated_tags:
                    tag_clean = tag.strip()
                    if not tag_clean or not isinstance(tag_clean, str): print_warning(f"Skipping invalid tag value: '{tag}'", indent=4); continue
                    try:
                        print_info(f"Adding tag word by word: '{tag_clean}'", indent=4)
                        words = tag_clean.split(' '); num_words = len(words)
                        for index, word in enumerate(words):
                            if not word: continue
                            tags_input.send_keys(word); mimic_human_action_delay(0.03, 0.08)
                            if index < num_words - 1: tags_input.send_keys(" "); mimic_human_action_delay(0.03, 0.08)
                        mimic_human_action_delay(0.1, 0.2)
                        tags_input.send_keys(Keys.ENTER); mimic_human_action_delay(0.2, 0.4) # Wait slightly longer after Enter
                        tags_added_count += 1
                    except Exception as tag_e:
                        print_warning(f"Error adding tag '{tag_clean[:30]}...': {tag_e}. Continuing...", indent=4)
                        log_error_to_file(f"Warning: Error adding tag '{tag_clean}' for video {video_index}: {tag_e}")
                        try: tags_input.send_keys(Keys.ESCAPE); time.sleep(0.2) # Try to recover focus
                        except Exception as recovery_err: print_warning(f"Minor issue during tag error recovery: {recovery_err}", indent=5)
                print_success(f"Attempted to add {tags_added_count}/{len(validated_tags)} tags.", indent=3)
            else: print_info("No valid tags found in metadata to add.", indent=3)
        except TimeoutException: print_error(f"Timeout finding Tags input field. Tags will likely be missing.", indent=3); log_error_to_file(f"ERROR: Timeout finding Tags input for {video_index}. Tags not added.");
        except JavascriptException as js_e: print_error(f"JavaScript error during tag input handling (likely scrolling): {js_e}", indent=3, include_traceback=False); log_error_to_file(f"Error (JS) adding tags {video_index}: {js_e}", include_traceback=True)
        except Exception as e: print_error(f"Unexpected error adding tags: {e}", indent=3, include_traceback=False); log_error_to_file(f"Error adding tags {video_index}: {e}", include_traceback=True)
        # --- End Tags ---

        # --- Select Category (with Dynamic Suggestion Fallback) ---
        suggested_category = metadata.get('suggested_category') # Get suggested category from loaded metadata
        category_to_use = upload_category # Default to category from config.txt

        if suggested_category:
            print_info(f"Attempting to use suggested category: '{suggested_category}'", indent=2)
            category_to_use = suggested_category # Prioritize suggested category
        else:
            print_info(f"No suggested category found in metadata. Using default from config: '{upload_category}'", indent=2)

        category_selected_successfully = False
        try:
            # Use category_to_use (either suggested or default) in the XPaths
            print_info(f"Selecting category: '{category_to_use}'...", indent=2)
            category_dropdown_xpath_options = [
                "//ytcp-form-select[@id='category']//ytcp-dropdown-trigger",
                "//ytcp-form-select[@id='category']//div[@class='left-container style-scope ytcp-dropdown-trigger']",
            ]
            # Dynamically create option XPaths based on the category we are trying to use
            category_option_xpath_options = [
                f"//yt-formatted-string[text()='{category_to_use}']",
                f"//yt-formatted-string[normalize-space()='{category_to_use}']",
                f"//ytcp-ve[@class='ytcp-form-select-options' and not(@hidden)]//yt-formatted-string[normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='{category_to_use.lower()}']"
            ]

            category_dropdown = None
            for i, xpath in enumerate(category_dropdown_xpath_options):
                try:
                    print_info(f"Trying category dropdown XPath {i+1}: {xpath}", indent=3)
                    category_dropdown = wait_short.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    print_success(f"Found category dropdown with XPath {i+1}", indent=3); break
                except Exception:
                    print_warning(f"Category dropdown XPath {i+1} failed, trying next...", indent=3);
                    if i == len(category_dropdown_xpath_options) - 1:
                        print_error("All category dropdown XPaths failed.", indent=3);
                        raise TimeoutException("Failed to find category dropdown with any XPath.") # Re-raise if dropdown fails

            # Attempt to click dropdown and select the category_to_use
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", category_dropdown); time.sleep(0.3)
            category_dropdown.click(); mimic_human_action_delay(0.4, 0.8);
            category_option = None
            for i_opt, xpath_opt in enumerate(category_option_xpath_options):
                try:
                    print_info(f"Trying category option XPath {i_opt+1}: {xpath_opt}", indent=3);
                    category_option = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_opt)))
                    print_success(f"Found category option '{category_to_use}' with XPath {i_opt+1}", indent=3); break
                except Exception:
                    print_warning(f"Category option XPath {i_opt+1} failed, trying next...", indent=3);
                    if i_opt == len(category_option_xpath_options) - 1:
                        # Option not found, raise exception to trigger fallback
                        raise TimeoutException(f"Failed to find category option '{category_to_use}' with any XPath.")

            category_option.click(); mimic_human_action_delay(0.2, 0.4);
            print_success(f"Category '{category_to_use}' selected.", indent=3)
            category_selected_successfully = True

        except Exception as category_err:
            print_error(f"Error selecting category '{category_to_use}': {category_err}. ", indent=3, include_traceback=False)
            log_error_to_file(f"Error selecting category '{category_to_use}' for {video_index}: {category_err}", include_traceback=True)

            # --- Fallback Logic ---
            if suggested_category and category_to_use == suggested_category:
                # The suggested category failed, now try the default from config
                print_warning(f"Attempting fallback to default category: '{upload_category}'", indent=3)
                category_to_use = upload_category # Set category to use to the default
                try:
                    # Re-attempt finding the dropdown (might be closed after error)
                    # NOTE: Assuming dropdown finder logic is robust enough, otherwise repeat it here.
                    # If dropdown click failed initially, this might also fail.
                    # We need to find the *option* for the default category now.
                    category_option_xpath_options_fallback = [
                        f"//yt-formatted-string[text()='{upload_category}']",
                        f"//yt-formatted-string[normalize-space()='{upload_category}']",
                        f"//ytcp-ve[@class='ytcp-form-select-options' and not(@hidden)]//yt-formatted-string[normalize-space(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'))='{upload_category.lower()}']"
                     ]
                    category_option_fallback = None
                    for i_opt, xpath_opt in enumerate(category_option_xpath_options_fallback):
                         try:
                             print_info(f"Trying category option XPath (fallback) {i_opt+1}: {xpath_opt}", indent=3);
                             category_option_fallback = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_opt)))
                             print_success(f"Found category option '{upload_category}' (fallback) with XPath {i_opt+1}", indent=3); break
                         except Exception:
                             print_warning(f"Category option XPath (fallback) {i_opt+1} failed...", indent=3);
                             if i_opt == len(category_option_xpath_options_fallback) - 1:
                                 raise TimeoutException(f"Failed to find fallback category option '{upload_category}'.")

                    category_option_fallback.click(); mimic_human_action_delay(0.2, 0.4);
                    print_success(f"Successfully selected default category '{upload_category}' after suggested failed.", indent=3)
                    category_selected_successfully = True # Mark success after fallback

                except Exception as fallback_err:
                     print_error(f"Fallback to default category '{upload_category}' also failed: {fallback_err}", indent=3, include_traceback=False)
                     log_error_to_file(f"Error selecting fallback category '{upload_category}' for {video_index}: {fallback_err}", include_traceback=True)
                     # Continue without guaranteeing category selection, or maybe raise error? For now, continue.
                     print_warning("Proceeding with upload, but category may not be set correctly.", indent=3)
            else:
                 # The default category failed initially, or it was an unexpected error
                 print_warning(f"Could not select the primary category '{category_to_use}'. Proceeding, but category may be incorrect.", indent=3)
                 # Continue without guaranteeing category selection


        # --- End Category Selection ---


        # --- Click through "Next" buttons ---
        print_info("Proceeding through 'Next' steps (Checks, Elements, Visibility)...", indent=2)

        # Add a longer pause before attempting the first "Next" click
        # This gives the UI more time to settle after filling metadata
        print_info("Pausing briefly before clicking first 'Next'...", indent=2)
        time.sleep(2.5)  # Increase pause duration to allow UI to stabilize

        num_next_clicks = 3; max_retries_next = 3; retry_delay_seconds_next = 5
        short_timeout_duration = 25
        next_button_xpath_options = [
            "//ytcp-button[@id='next-button' and not(@disabled)]", # Preferred ID, not disabled
            "//div[@id='dialog-buttons']//ytcp-button[contains(., 'Next') and not(@disabled)]", # Within dialog buttons div
        ]
        for i in range(num_next_clicks):
            step_name = ["Details", "Checks", "Visibility"][i]
            print_info(f"Attempting to click 'Next' for Step {i+1} ({step_name})...", indent=3); next_button_clicked = False; retry_count = 0
            while retry_count < max_retries_next and not next_button_clicked:
                next_button = None
                for i_xpath, xpath in enumerate(next_button_xpath_options):
                    try:
                        print_info(f"Trying Next button XPath {i_xpath+1}: {xpath}", indent=4)
                        # Use slightly shorter timeout per attempt within the retry loop
                        next_button = WebDriverWait(driver, short_timeout_duration // (len(next_button_xpath_options) * (retry_count + 1)) + 2 ).until(
                            EC.element_to_be_clickable((By.XPATH, xpath))
                        )
                        print_success(f"Found Next button with XPath {i_xpath+1}", indent=4); break
                    except Exception: print_warning(f"Next button XPath {i_xpath+1} failed...", indent=4);
                if next_button:
                    try:
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_button); mimic_human_action_delay(0.2, 0.5)
                        next_button.click(); next_button_clicked = True
                        print_success(f"'Next' button (Step {i+1} - {step_name}) clicked successfully.", indent=4)
                        # Wait longer after 'Checks' step
                        wait_after_next = 3.0 if i == 1 else 2.0
                        print_info(f"Waiting {wait_after_next}s after clicking Next...", indent=5)
                        time.sleep(wait_after_next) # Use time.sleep for fixed delay
                        break # Exit retry loop
                    except ElementClickInterceptedException:
                        print_warning(f"'Next' button (Step {i+1}) intercepted. Trying JS click...", indent=4)
                        try:
                            # Try JavaScript click when intercepted
                            driver.execute_script("arguments[0].click();", next_button)
                            next_button_clicked = True
                            print_success(f"'Next' button (Step {i+1} - {step_name}) clicked successfully (JS Click).", indent=4)
                            # Wait after click
                            wait_after_next = 3.0 if i == 1 else 2.0
                            print_info(f"Waiting {wait_after_next}s after clicking Next...", indent=5)
                            time.sleep(wait_after_next)
                            break # Exit retry loop
                        except Exception as js_e:
                            print_warning(f"JS click also failed for 'Next' (Step {i+1}): {js_e}", indent=4)
                            # Try scrolling and waiting before next attempt
                            driver.execute_script("window.scrollBy(0, 150);") # Scroll down slightly
                            time.sleep(1) # Wait a bit longer after failed JS click
                    except Exception as e:
                        print_error(f"Unexpected error clicking found 'Next' button (Step {i+1}): {e}", indent=4, include_traceback=False)
                        log_error_to_file(f"Error clicking Next {i+1} ({step_name}) for {video_index}: {e}", include_traceback=True)
                        # Don't return False immediately, let the retry loop continue
                if not next_button_clicked:
                    print_warning(f"'Next' button (Step {i+1}) not clickable or found. Retry {retry_count+1}/{max_retries_next} in {retry_delay_seconds_next}s...", indent=4)
                    time.sleep(retry_delay_seconds_next); retry_count += 1
            if not next_button_clicked:
                msg = f"Failed to find or click 'Next' button (Step {i+1} - {step_name}) after {max_retries_next} attempts."
                print_error(msg, indent=3)
                log_error_to_file(f"ERROR: {msg}", error_type="next_button", step=f"Next_{i+1}", video_index=video_index)
                return None  # Return None instead of False to clearly indicate failure
        # --- End Next Buttons ---

        # --- Handle Visibility (Publish or Schedule) ---
        print_info("Setting visibility...", indent=2)
        visibility_wait = WebDriverWait(driver, 30) # Wait for visibility options

        if publish_now:
            # --- Publish Now Logic ---
            try:
                print_info("Selecting 'Public' visibility...", indent=3)
                public_radio_xpath_options = [
                    "//tp-yt-paper-radio-button[@name='PUBLIC']", # Preferred name attribute
                    "//div[@id='visibility-list']//tp-yt-paper-radio-button[contains(., 'Public')]", # Text within container
                ]
                public_radio = None
                for i, xpath in enumerate(public_radio_xpath_options):
                    try:
                        print_info(f"Trying Public radio XPath {i+1}: {xpath}", indent=4)
                        public_radio = visibility_wait.until(EC.element_to_be_clickable((By.XPATH, xpath)), message=f"Timeout finding Public Radio XPath {i+1}")
                        print_success(f"Found Public radio with XPath {i+1}", indent=4); break
                    except Exception: print_warning(f"Public radio XPath {i+1} failed, trying next...", indent=4);
                    if i == len(public_radio_xpath_options) - 1: raise TimeoutException("Failed to find Public radio button")

                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", public_radio); time.sleep(0.5)
                try: print_info("Attempting normal click on Public radio...", indent=4); public_radio.click()
                except Exception as click_e: print_warning(f"Normal click failed: {click_e}, trying JavaScript click...", indent=4); driver.execute_script("arguments[0].click();", public_radio)
                print_success("'Public' selected.", indent=4); mimic_human_action_delay(0.5, 1.0)

                done_button_xpath_options = [
                    "//ytcp-button[@id='done-button' and not(@disabled)]", # Preferred ID, not disabled
                    "//div[@id='dialog-buttons']//ytcp-button[contains(., 'Publish') and not(@disabled)]", # Text within container
                ]
                print_info("Looking for Publish button...", indent=3); publish_button = None
                for i, xpath in enumerate(done_button_xpath_options):
                    try:
                        print_info(f"Trying Publish button XPath {i+1}: {xpath}", indent=4)
                        publish_button = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath)), message=f"Timeout finding Publish Button XPath {i+1}")
                        print_success(f"Found Publish button with XPath {i+1}", indent=4); break
                    except Exception: print_warning(f"Publish button XPath {i+1} failed, trying next...", indent=4);
                    if i == len(done_button_xpath_options) - 1: raise TimeoutException("Failed to find Publish button")

                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", publish_button); time.sleep(0.5)
                print_info("Attempting to click Publish button...", indent=4)
                try: publish_button.click(); print_info("Publish button clicked (normal click).", indent=4)
                except Exception as click_e: print_warning(f"Normal click failed: {click_e}, trying JavaScript click...", indent=4); driver.execute_script("arguments[0].click();", publish_button); print_info("Publish button clicked (JS click).", indent=4)

                # --- NEW/MODIFIED Confirmation & ID Capture Logic ---
                # Wait specifically for the success dialog or a key element within it (like the share URL)
                # More specific selectors for the share dialog and URL elements
                success_dialog_xpath = "//tp-yt-iron-overlay-backdrop[@class='opened']" # XPath for the opened dialog backdrop
                share_url_element_xpath = "//a[contains(@href, 'youtube.com/shorts/') or contains(@href, 'youtu.be/')]" # XPath for any share URL link
                copy_button_xpath = "//ytcp-icon-button[@class='right style-scope ytcp-video-share-dialog']//tp-yt-iron-icon" # Copy button in share dialog

                print_info("Waiting for upload success dialog and Share URL element...", indent=3)
                try:
                    # Wait for the success dialog to be present first
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, success_dialog_xpath)), message="Timeout waiting for upload success dialog.")
                    # Then wait for the share URL element within that dialog
                    share_url_element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, share_url_element_xpath)), message="Timeout waiting for Share URL element in success dialog.")

                    # Also check for the copy button as an additional confirmation that we're in the share dialog
                    try:
                        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, copy_button_xpath)), message="Copy button not found in share dialog.")
                        print_info("Share dialog copy button found - confirmed in share dialog.", indent=4)
                    except Exception:
                        print_info("Copy button not found, but proceeding with URL extraction anyway.", indent=4)

                    # --- Extract the Video ID ---
                    share_url = share_url_element.get_attribute('href')
                    print_info(f"Captured Share URL: {share_url}", indent=4)

                    if share_url:
                        # The share URL could be in different formats:
                        # - https://youtu.be/VIDEO_ID
                        # - https://youtube.com/shorts/VIDEO_ID
                        # - https://youtube.com/watch?v=VIDEO_ID
                        match = re.search(r"(?:youtu\.be/|youtube\.com/(?:shorts/|watch\?v=))([^?&]+)", share_url)
                        if match:
                            youtube_video_id = match.group(1)
                            print_success(f"Parsed YouTube Video ID: {youtube_video_id}", indent=4)
                            upload_successful = True # <<< Set success *here* after capturing ID
                        else:
                            print_warning(f"Could not parse YouTube Video ID from Share URL: {share_url}", indent=4)
                            upload_successful = False # Mark as failed if ID not parsed
                    else:
                         print_warning("Share URL element found but href attribute is empty.", indent=4)
                         upload_successful = False

                    # --- Close the Success Dialog AFTER capturing ID ---
                    print_info("Looking for dialog close button...", indent=4)
                    close_button_xpath = "//ytcp-button[@id='close-button'] | //ytcp-icon-button[@aria-label='Close']" # Standard dialog close buttons
                    try:
                         # Use a short wait here, the button should be visible immediately
                         close_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, close_button_xpath)), message="Timeout finding dialog close button.")
                         driver.execute_script("arguments[0].scrollIntoView({block:'center'});", close_button); time.sleep(0.2)
                         close_button.click(); mimic_human_action_delay(0.5, 1.0)
                         print_success("Upload success dialog closed.", indent=4)
                    except TimeoutException as close_err:
                         print_warning(f"Timeout finding dialog close button. Dialog may remain open. {close_err}", indent=4)
                    except Exception as close_err:
                         print_warning(f"Error clicking dialog close button: {close_err}. Dialog may remain open.", indent=4)

                except TimeoutException as e:
                    msg = f"Timeout waiting for upload success dialog or Share URL element: {e}."
                    print_error(msg, indent=3); log_error_to_file(f"Error: Timeout confirming Publish/getting ID for {video_index}: {e}")
                    upload_successful = False # Ensure false state on timeout here
                    youtube_video_id = None # Ensure ID is None

                except Exception as e:
                    msg = f"Unexpected error during Publish confirmation/ID capture: {e}"
                    print_error(msg, indent=3, include_traceback=True); log_error_to_file(f"Error: Unexpected error during Publish confirmation/ID capture for {video_index}: {e}", include_traceback=True)
                    upload_successful = False # Ensure false state on error here
                    youtube_video_id = None # Ensure ID is None
                # --- END NEW/MODIFIED Confirmation & ID Capture Logic ---

            except TimeoutException as e: msg = f"Timeout during 'Publish Now' step: {e}."; print_error(msg, indent=3); return False
            except Exception as e: msg = f"Unexpected error during 'Publish Now': {e}"; print_error(msg, indent=3, include_traceback=True); return False
            # --- End Publish Now ---

        else: # Schedule
            # --- Schedule Logic (Using OLD Method) ---
            if not schedule_time: print_error("Schedule time was not provided.", indent=3); return False
            try:
                schedule_radio_xpath_options = [
                    "//ytcp-icon-button[@id='second-container-expand-button']", # User provided (mistaken for expand)
                    "//tp-yt-paper-radio-button[@name='SCHEDULE']", # Original best guess
                    "//div[contains(@id, 'second-container')]//ytcp-icon-button", # User provided (mistaken for expand)
                    # Add any other potential Schedule radio XPaths here if needed
                ]
                print_info(f"Selecting 'Schedule' for {schedule_time:%Y-%m-%d %H:%M:%S}...", indent=3)
                schedule_radio = None
                for i, xpath in enumerate(schedule_radio_xpath_options):
                    try:
                        print_info(f"Trying Schedule radio XPath {i+1}: {xpath}", indent=4)
                        schedule_radio = visibility_wait.until(EC.element_to_be_clickable((By.XPATH, xpath)), message=f"Timeout finding Schedule Radio XPath {i+1}")
                        print_success(f"Found Schedule radio with XPath {i+1}", indent=4); break
                    except Exception: print_warning(f"Schedule radio XPath {i+1} failed, trying next...", indent=4);
                    if i == len(schedule_radio_xpath_options) - 1: raise TimeoutException("Failed to find Schedule radio button with any XPath.")

                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", schedule_radio); time.sleep(0.3)
                if schedule_radio.get_attribute("aria-checked") != "true":
                    try: schedule_radio.click()
                    except: driver.execute_script("arguments[0].click();", schedule_radio) # JS fallback click
                    mimic_human_action_delay(0.5, 1.0); print_success("Selected 'Schedule'.", indent=4)
                else: print_info("'Schedule' is already selected.", indent=4)

                # --- Set Date (Old Flow + Fallbacks) ---
                try:
                    schedule_date_trigger_xpath_options = [
                        "//ytcp-text-dropdown-trigger[@id='datepicker-trigger']", # Preferred ID
                        "//ytcp-dropdown-trigger[contains(@class, 'ytcp-date-picker')]", # Class based
                    ]
                    schedule_date_trigger = None
                    for i, xpath in enumerate(schedule_date_trigger_xpath_options):
                         try:
                             print_info(f"Trying Date trigger XPath {i+1}: {xpath}", indent=3)
                             schedule_date_trigger = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)), message=f"Timeout finding Date Trigger XPath {i+1}")
                             print_success(f"Found Date trigger with XPath {i+1}", indent=3); break
                         except Exception: print_warning(f"Date trigger XPath {i+1} failed...", indent=3);
                         if i == len(schedule_date_trigger_xpath_options) - 1: raise TimeoutException("Failed to find Date trigger")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", schedule_date_trigger); mimic_human_action_delay(0.2, 0.5)
                    schedule_date_trigger.click(); print_success("Date dropdown trigger clicked.", indent=4); mimic_human_action_delay(0.5, 1.0)

                    print_info("Calling OLD function to select date in calendar...", indent=3)
                    if not select_date_in_calendar(driver, schedule_time): raise Exception("Failed to set schedule date using old method") # Explicit failure
                    print_success("Date selection function completed.", indent=4); mimic_human_action_delay(0.5, 1.0)
                except Exception as date_err: msg = f"Error during date selection process (old method+fallbacks): {date_err}"; print_error(msg, indent=3, include_traceback=True); return False # Log full traceback for date errors
                # --- End Set Date ---

                # --- Set Time (Old Flow + Fallbacks) ---
                try:
                    time_input_xpath_options = [
                        "//ytcp-form-input-container[@id='time-of-day-container']//input[@class='style-scope tp-yt-paper-input']", # Original
                        "//ytcp-form-input-container[@id='time-of-day-container']//input[contains(@aria-label, 'Time')]", # Aria label
                    ]
                    print_info("Setting schedule time using OLD method (.clear(), AM/PM + Fallbacks)...", indent=3)
                    schedule_time_input = None
                    for i, xpath in enumerate(time_input_xpath_options):
                        try:
                            print_info(f"Trying Time input XPath {i+1}: {xpath}", indent=3)
                            schedule_time_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)), message=f"Timeout finding Time Input XPath {i+1}")
                            print_success(f"Found Time input with XPath {i+1}", indent=3); break
                        except Exception: print_warning(f"Time input XPath {i+1} failed...", indent=3);
                        if i == len(time_input_xpath_options) - 1: raise TimeoutException("Failed to find Time input field")

                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", schedule_time_input); mimic_human_action_delay(0.2, 0.5)
                    try: WebDriverWait(driver, 5).until(EC.element_to_be_clickable(schedule_time_input)); schedule_time_input.click()
                    except: print_warning("Could not click time input for focus, trying JS click...", indent=4);
                    try: driver.execute_script("arguments[0].click();", schedule_time_input);
                    except: print_warning("JS click failed.", indent=4)

                    mimic_human_action_delay(0.2, 0.5); print_info("Clearing time input field using .clear()...", indent=4)
                    schedule_time_input.clear(); mimic_human_action_delay(0.2, 0.5)
                    # Use 24-hour format (HH:MM) which is generally more reliable for input
                    time_str = schedule_time.strftime("%H:%M")
                    print_info(f"Formatted time string (24-hour): '{time_str}'", indent=4)
                    schedule_time_input.send_keys(time_str); mimic_human_action_delay(0.4, 0.7)
                    print_info("Sending Enter key for time input...", indent=4)
                    schedule_time_input.send_keys(Keys.ENTER); mimic_human_action_delay(0.5, 1.0)
                    print_success("Schedule time set using OLD method.", indent=4)
                except Exception as time_err: msg = f"Error during time input process (old method+fallbacks): {time_err}"; print_error(msg, indent=3, include_traceback=True); return False
                # --- End Set Time ---

                # --- Click Final Schedule Button ---
                try:
                    done_button_xpath_options = [
                        "//ytcp-button[@id='done-button' and not(@disabled)]", # Preferred ID
                         "//div[@id='dialog-buttons']//ytcp-button[contains(., 'Schedule') and not(@disabled)]", # Text based fallback
                    ]
                    print_info("Finding final 'Schedule' (Done) button...", indent=3); schedule_button = None
                    for i, xpath in enumerate(done_button_xpath_options):
                         try:
                             print_info(f"Trying Done/Schedule button XPath {i+1}: {xpath}", indent=3)
                             schedule_button = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath)), message=f"Timeout finding Schedule Button XPath {i+1}")
                             print_success(f"Found Done/Schedule button with XPath {i+1}", indent=3); break
                         except Exception: print_warning(f"Done/Schedule button XPath {i+1} failed...", indent=3);
                         if i == len(done_button_xpath_options) - 1: raise TimeoutException("Failed to find Done/Schedule button")
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", schedule_button); mimic_human_action_delay(0.2, 0.5)
                    print_info("Attempting to click final 'Schedule' button...", indent=4)
                    try: schedule_button.click(); print_info("Schedule button clicked (normal click).", indent=4)
                    except: print_warning("Normal click intercepted/failed, trying JS click...", indent=4); driver.execute_script("arguments[0].click();", schedule_button); print_info("Schedule button clicked (JS click).", indent=4)

                    # --- NEW/MODIFIED Confirmation & ID Capture Logic (Similar to Publish) ---
                    # Wait specifically for the success dialog or a key element within it (like the share URL)
                    # More specific selectors for the share dialog and URL elements
                    success_dialog_xpath = "//tp-yt-iron-overlay-backdrop[@class='opened']" # XPath for the opened dialog backdrop
                    share_url_element_xpath = "//a[contains(@href, 'youtube.com/shorts/') or contains(@href, 'youtu.be/')]" # XPath for any share URL link
                    copy_button_xpath = "//ytcp-icon-button[@class='right style-scope ytcp-video-share-dialog']//tp-yt-iron-icon" # Copy button in share dialog

                    print_info("Waiting for schedule success dialog and Share URL element...", indent=3)
                    try:
                        # Wait for the success dialog to be present first
                        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, success_dialog_xpath)), message="Timeout waiting for schedule success dialog.")
                        # Then wait for the share URL element within that dialog
                        share_url_element = WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, share_url_element_xpath)), message="Timeout waiting for Share URL element in success dialog.")

                        # Also check for the copy button as an additional confirmation that we're in the share dialog
                        try:
                            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, copy_button_xpath)), message="Copy button not found in share dialog.")
                            print_info("Share dialog copy button found - confirmed in share dialog.", indent=4)
                        except Exception:
                            print_info("Copy button not found, but proceeding with URL extraction anyway.", indent=4)

                        # --- Extract the Video ID ---
                        share_url = share_url_element.get_attribute('href')
                        print_info(f"Captured Share URL: {share_url}", indent=4)

                        if share_url:
                            # The share URL could be in different formats:
                            # - https://youtu.be/VIDEO_ID
                            # - https://youtube.com/shorts/VIDEO_ID
                            # - https://youtube.com/watch?v=VIDEO_ID
                            match = re.search(r"(?:youtu\.be/|youtube\.com/(?:shorts/|watch\?v=))([^?&]+)", share_url)
                            if match:
                                youtube_video_id = match.group(1)
                                print_success(f"Parsed YouTube Video ID: {youtube_video_id}", indent=4)
                                upload_successful = True # <<< Set success *here* after capturing ID
                            else:
                                print_warning(f"Could not parse YouTube Video ID from Share URL: {share_url}", indent=4)
                                upload_successful = False
                        else:
                             print_warning("Share URL element found but href attribute is empty.", indent=4)
                             upload_successful = False


                        # --- Close the Success Dialog AFTER capturing ID ---
                        print_info("Looking for dialog close button...", indent=4)
                        close_button_xpath = "//ytcp-button[@id='close-button'] | //ytcp-icon-button[@aria-label='Close']"
                        try:
                             close_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, close_button_xpath)), message="Timeout finding dialog close button.")
                             driver.execute_script("arguments[0].scrollIntoView({block:'center'});", close_button); time.sleep(0.2)
                             close_button.click(); mimic_human_action_delay(0.5, 1.0)
                             print_success("Schedule success dialog closed.", indent=4)
                        except TimeoutException as close_err:
                             print_warning(f"Timeout finding dialog close button. Dialog may remain open. {close_err}", indent=4)
                        except Exception as close_err:
                             print_warning(f"Error clicking dialog close button: {close_err}. Dialog may remain open.", indent=4)

                    except TimeoutException as e:
                        msg = f"Timeout waiting for schedule success dialog or Share URL element: {e}."
                        print_error(msg, indent=3); log_error_to_file(f"Error: Timeout confirming Schedule/getting ID for {video_index}: {e}")
                        upload_successful = False # Ensure false state
                        youtube_video_id = None # Ensure ID is None
                    except Exception as e:
                        msg = f"Unexpected error during Schedule confirmation/ID capture: {e}"
                        print_error(msg, indent=3, include_traceback=True); log_error_to_file(f"Error: Unexpected error during Schedule confirmation/ID capture for {video_index}: {e}", include_traceback=True)
                        upload_successful = False # Ensure false state
                        youtube_video_id = None # Ensure ID is None
                    # --- END NEW/MODIFIED Confirmation & ID Capture Logic ---

                except Exception as final_err: msg = f"Error clicking final 'Schedule' button: {final_err}"; print_error(msg, indent=3, include_traceback=True); return False
                # --- End Click Final Schedule Button ---

            except Exception as e: msg = f"Unexpected error during 'Schedule' block: {e}"; print_error(msg, indent=3, include_traceback=True); return False
            # --- End Schedule Logic ---

    # --- General Exception Handling ---
    except (NoSuchWindowException, InvalidSessionIdException) as sess_err: msg = f"Browser session lost during upload for {video_index}: {sess_err}"; print_error(msg, indent=1); raise # Re-raise critical errors
    except TimeoutException as e: msg = f"Timeout waiting for essential element during upload for {video_index}: {e}"; print_error(msg, indent=1); return False
    except Exception as e: msg = f"Major unexpected error during upload steps for {video_index}: {e}"; print_error(msg, indent=1, include_traceback=True); return False
    finally:
        print_info(f"--- Finished Upload Attempt for Video Index: {video_index} ---", indent=1)

    # Return the video ID if successful, None otherwise
    return youtube_video_id


# --- delete_uploaded_files function ---
def delete_uploaded_files(video_file: str, metadata_file_path: str) -> bool:
    """Deletes the video and its corresponding metadata JSON file."""
    print_info(f"Attempting cleanup for: {os.path.basename(video_file)}", indent=1)
    deleted_video = False; deleted_metadata = False
    try:
        if os.path.exists(video_file): os.remove(video_file); print_success(f"Deleted video file: {os.path.basename(video_file)}", indent=2); deleted_video = True
        else: print_info(f"Video file not found (already deleted?): {os.path.basename(video_file)}", indent=2); deleted_video = True # Treat missing as success for overall cleanup
    except OSError as e: print_error(f"OS Error deleting video file {os.path.basename(video_file)}: {e}", indent=2)
    except Exception as e: print_error(f"Unexpected error deleting video file {os.path.basename(video_file)}: {e}", indent=2)
    try:
        if os.path.exists(metadata_file_path): os.remove(metadata_file_path); print_success(f"Deleted metadata file: {os.path.basename(metadata_file_path)}", indent=2); deleted_metadata = True
        else: print_info(f"Metadata file not found (already deleted?): {os.path.basename(metadata_file_path)}", indent=2); deleted_metadata = True # Treat missing as success
    except OSError as e: print_error(f"OS Error deleting metadata file {os.path.basename(metadata_file_path)}: {e}", indent=2)
    except Exception as e: print_error(f"Unexpected error deleting metadata file {os.path.basename(metadata_file_path)}: {e}", indent=2)
    success = deleted_video and deleted_metadata
    if not success: print_warning("Cleanup incomplete. Some files may remain.", indent=1)
    return success

# --- FFmpeg Recording Helper Functions ---
def start_recording(video_index: str, ffmpeg_cmd_path: str, driver: webdriver.Firefox) -> Optional[Tuple[subprocess.Popen, str]]:
    """Starts the FFmpeg screen recording, attempting to target the Firefox window."""
    global _current_recording_process, _current_recording_filename # Use global vars

    if not driver:
        print_error("Cannot start window recording: Selenium driver is not valid.", indent=3)
        return None

    print_info(f"Attempting to start debug recording for video index: {video_index}", indent=2)
    try:
        os.makedirs(DEBUG_RECORDING_FOLDER, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = os.path.join(DEBUG_RECORDING_FOLDER, f"recording_video_{video_index}_{timestamp}.mp4")

        cmd = [ffmpeg_cmd_path, '-y', '-loglevel', 'error', '-f']
        system = platform.system()

        if system == "Windows":
             cmd.extend(['gdigrab']) # Base input format for Windows
             try:
                window_title = driver.title
                if not window_title or len(window_title) > 100: # Avoid overly long/problematic titles
                    print_warning(f"Firefox window title problematic ('{window_title[:50]}...'). Falling back to desktop capture.", indent=4)
                    cmd.extend(['-i', 'desktop'])
                else:
                    print_info(f"Targeting Firefox window with title: '{window_title}'", indent=4)
                    cmd.extend(['-i', f'title={window_title}'])
                    time.sleep(0.3) # Small delay might help FFmpeg latch on
             except WebDriverException as e:
                 print_error(f"Could not get Firefox window title via Selenium: {e}. Falling back to desktop capture.", indent=4)
                 cmd.extend(['-i', 'desktop'])
             except Exception as e:
                 print_error(f"Unexpected error getting window title: {e}. Falling back to desktop capture.", indent=4)
                 cmd.extend(['-i', 'desktop'])

             # Resolve local path if needed
             if not os.path.isabs(ffmpeg_cmd_path) and not any(os.path.exists(os.path.join(p, ffmpeg_cmd_path)) for p in os.environ.get("PATH", "").split(os.pathsep)):
                  local_path = os.path.join(script_directory, ffmpeg_cmd_path)
                  if os.path.exists(local_path + ".exe"): cmd[0] = local_path + ".exe" # Check with .exe too
                  elif os.path.exists(local_path): cmd[0] = local_path


        elif system == "Linux":
             # Basic X11 capture (may need adjustment based on environment/Wayland)
             # Getting specific window ID on Linux is complex and varies; desktop capture is simpler fallback
             print_warning("Linux detected: Using full desktop capture for recording (specific window capture not implemented). Ensure 'x11grab' is supported.", indent=4)
             display = os.environ.get('DISPLAY')
             if not display:
                 print_error("Linux: DISPLAY environment variable not set. Cannot use x11grab.", indent=3)
                 return None
             cmd.extend(['x11grab', '-i', display])
             # Add framerate before input for x11grab
             cmd.insert(-2, '-r')
             cmd.insert(-2, '15') # Framerate before input


        elif system == "Darwin": # macOS
             # Basic AVFoundation capture (may need permissions)
             print_warning("macOS detected: Using full desktop capture via AVFoundation. Ensure screen recording permission is granted.", indent=4)
             cmd.extend(['avfoundation', '-i', '1:none']) # Capture screen 1, no audio input
             # Add framerate before input for avfoundation
             cmd.insert(-2, '-r')
             cmd.insert(-2, '15')


        else:
             print_error(f"Unsupported OS for recording: {system}", indent=3)
             return None

        # Check if input was actually added
        if len(cmd) <= 5:
             print_error("Failed to determine input method for FFmpeg command.", indent=3)
             return None

        print_info(f"Recording from: {' '.join(cmd[4:6])}", indent=3)
        print_info(f"Recording to: {output_filename}", indent=3)

        # Common Output settings
        cmd.extend([
            '-r', '15', # Output framerate (redundant for some inputs, but safe)
            '-c:v', 'libx264',
            '-preset', 'ultrafast',
            '-crf', '28', # Visual quality vs size trade-off
            '-pix_fmt', 'yuv420p', # Ensures compatibility
            output_filename
        ])

        print_info(f"Full FFmpeg command list: {cmd}", indent=3)

        creationflags = 0
        if system == "Windows": creationflags = subprocess.CREATE_NO_WINDOW

        process = subprocess.Popen(
            cmd,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            creationflags=creationflags
        )

        time.sleep(2.0) # Allow FFmpeg to initialize

        if process.poll() is not None:
            stderr_output = ""
            try: stderr_output = process.stderr.read().decode('utf-8', errors='ignore')
            except Exception as read_err: print_warning(f"Could not read FFmpeg stderr on immediate exit: {read_err}", indent=4)

            print_error(f"FFmpeg failed to start or exited immediately (return code: {process.returncode}). Error output:", indent=3)
            if stderr_output:
                 print(f"{Fore.RED}{stderr_output}{Style.RESET_ALL}")
                 if "Could not find window" in stderr_output: print_warning("This might indicate the window title was not found or changed.", indent=4)
            else: print_info("No stderr output captured.", indent=4)
            log_error_to_file(f"FFmpeg failed to start for video {video_index}. Command: {' '.join(cmd)}\nReturn Code: {process.returncode}\nStderr: {stderr_output}")
            return None

        print_success("FFmpeg recording process started.", indent=3)
        _current_recording_process = process
        _current_recording_filename = output_filename
        return process, output_filename

    except FileNotFoundError:
        print_error(f"FFmpeg executable not found at path specified or in system PATH: '{ffmpeg_cmd_path}'.", indent=3)
        print_error("Check FFMPEG_PATH in config.txt and ensure FFmpeg is installed correctly.", indent=3)
        log_error_to_file(f"FFmpeg not found at path: {ffmpeg_cmd_path}")
        return None
    except WebDriverException as e:
         print_error(f"Selenium error during recording setup (getting title?): {e}", indent=3, include_traceback=True)
         log_error_to_file(f"Selenium error during recording setup for video {video_index}: {e}", include_traceback=True)
         return None
    except OSError as e:
        print_error(f"OS Error starting FFmpeg recording: {e}", indent=3, include_traceback=True)
        log_error_to_file(f"OS Error starting FFmpeg recording for video {video_index}: {e}", include_traceback=True)
        return None
    except Exception as e:
        print_error(f"Unexpected error starting FFmpeg recording: {e}", indent=3, include_traceback=True)
        log_error_to_file(f"Unexpected error starting FFmpeg recording for video {video_index}: {e}", include_traceback=True)
        return None

def stop_recording(process: Optional[subprocess.Popen], filename: Optional[str], keep_file: bool):
    """Stops the FFmpeg recording process and handles the output file."""
    global _current_recording_process, _current_recording_filename

    if not process or not filename:
        print_info("No active recording process/filename to stop.", indent=2)
        return

    print_info(f"Stopping recording process (PID: {process.pid}). Keep file: {keep_file}", indent=2)

    if process.poll() is None: # Check if the process is still running
        print_info("Attempting graceful shutdown (sending 'q')...", indent=3)
        graceful_success = False; shutdown_timeout = 5

        try:
            if process.stdin:
                # Send 'q' for graceful exit
                if platform.system() == "Windows":
                    # On Windows, sending signal might be needed if 'q' fails often
                    # For now, stick to 'q' as primary method
                     process.stdin.write(b'q') # No newline needed typically for Windows pipe
                else:
                     process.stdin.write(b'q\n') # Newline often needed for Linux/macOS
                process.stdin.flush()
                process.stdin.close()
                print_info(f"Waiting up to {shutdown_timeout}s for FFmpeg to exit gracefully...", indent=4)
                process.wait(timeout=shutdown_timeout)
                graceful_success = True
                print_success("FFmpeg process exited gracefully.", indent=4)
            else: print_warning("Process stdin is not available, cannot send 'q'.", indent=4)
        except (subprocess.TimeoutExpired, OSError, BrokenPipeError, ValueError) as e:
            print_warning(f"Graceful shutdown failed or timed out: {e}. Attempting terminate/kill...", indent=4)
            graceful_success = False
        except Exception as e: print_warning(f"Unexpected error during graceful shutdown: {e}. Attempting terminate/kill...", indent=4); graceful_success = False

        # Force termination if graceful shutdown failed
        if not graceful_success and process.poll() is None:
            try:
                print_info("Terminating FFmpeg process...", indent=4)
                if platform.system() == "Windows":
                    # Use taskkill for potentially more robust termination on Windows
                    subprocess.run(['taskkill', '/F', '/PID', str(process.pid)], check=False, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                else:
                    # Use SIGINT (Ctrl+C equivalent) first, then SIGTERM, then SIGKILL
                    os.kill(process.pid, signal.SIGINT)
                    time.sleep(1) # Give it a moment to respond
                    if process.poll() is None:
                        os.kill(process.pid, signal.SIGTERM) # Standard terminate
                        time.sleep(1)
                        if process.poll() is None:
                            os.kill(process.pid, signal.SIGKILL) # Force kill

                process.wait(timeout=3) # Wait a bit after any terminate/kill attempt
                print_success("FFmpeg process terminated/killed.", indent=4)
            except (subprocess.TimeoutExpired, ProcessLookupError):
                print_warning("Process likely already exited after terminate/kill attempt.", indent=4)
            except Exception as term_e: print_error(f"Error during terminate/kill: {term_e}", indent=4)

    else: print_info(f"Recording process already terminated (Return code: {process.poll()}).", indent=3)

    # Handle the recording file
    if not keep_file:
        print_info(f"Deleting recording file: {filename}", indent=3)
        try:
            if os.path.exists(filename): os.remove(filename); print_success("Recording file deleted successfully.", indent=4)
            else: print_info("Recording file not found (already deleted?).", indent=4)
        except OSError as e: print_error(f"Error deleting recording file {filename}: {e}", indent=4); log_error_to_file(f"Error deleting recording file {filename}: {e}")
        except Exception as e: print_error(f"Unexpected error deleting recording file {filename}: {e}", indent=4, include_traceback=True); log_error_to_file(f"Unexpected error deleting recording file {filename}: {e}", include_traceback=True)
    else:
        if os.path.exists(filename): print_success(f"Keeping recording file: {filename}", indent=3)
        else: print_warning(f"Intended to keep recording, but file not found: {filename}", indent=3)


    # Clear global tracking variables *after* handling
    _current_recording_process = None
    _current_recording_filename = None
# --- End FFmpeg Recording Helper Functions ---


# --- Main Execution Logic ---
def main():
    global _current_recording_process, _current_recording_filename

    # Parse command-line arguments
    analyze_mode = False
    if len(sys.argv) > 1:
        if sys.argv[1].lower() in ["--analyze", "-a"]:
            analyze_mode = True
            print_section_header("Starting YouTube Uploader in Analysis Mode")
            if genai is None:
                print_error("Google Generative AI library not found. Cannot run analysis.")
                print_info("Install with: pip install google-generativeai")
                return

            # Check for API key
            gemini_api_key = ""
            try:
                with open(CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
                    for line in f:
                        if line.strip().startswith("GEMINI_API_KEY="):
                            gemini_api_key = line.strip().split("=", 1)[1].strip()
                            if gemini_api_key.startswith('"') and gemini_api_key.endswith('"'):
                                gemini_api_key = gemini_api_key[1:-1]
                            break
            except Exception as e:
                print_error(f"Error reading config file: {e}")

            if not gemini_api_key:
                print_error("GEMINI_API_KEY not found in config.txt. Cannot run analysis.")
                print_info("Add GEMINI_API_KEY=your_api_key to config.txt")
                return

            # Configure Gemini API
            try:
                genai.configure(api_key=gemini_api_key)
                print_success("Gemini API configured successfully.")
            except Exception as e:
                print_error(f"Failed to configure Gemini API: {e}")
                return

            # Run analysis
            print_info("Analyzing upload errors...")
            analysis = analyze_upload_errors_with_gemini()
            if analysis:
                print_success("Analysis completed successfully.")
                print_section_header("Analysis Results")
                print(analysis)
                print_info(f"\nFull analysis saved to {UPLOADER_ANALYSIS_LOG}")
            else:
                print_warning("Analysis could not be completed. See above messages for details.")
            return

    print_section_header("Starting YouTube Uploader Script"); start_time = time.time()
    driver = None; wb = None; downloaded_sheet = None; uploaded_sheet = None; excel_save_required = False

    # Initialize performance metrics for this run
    metrics = load_performance_metrics()
    metrics["total_uploads_attempted"] = 0
    metrics["total_uploads_successful"] = 0

    try:
        # --- Print Configuration ---
        print_section_header("Configuration Settings")
        print_config("Max Uploads per Run", max_uploads)
        print_config("Video Category", upload_category)
        print_config("Profile Path", profile_path_config if profile_path_config else f"{Style.DIM}Default{Style.RESET_ALL}")
        print_config("Metadata Folder", METADATA_FOLDER)
        print_config("Uploads Folder", UPLOAD_FOLDER)
        print_config("Excel Data File", EXCEL_FILE_PATH)
        print_config("Error Log File", ERROR_LOG_FILE)
        print_config("Desc Char Limit", cfg_desc_limit)
        print_config("Tag Char Limit", cfg_tag_limit)
        print_config("Total Tag Chars Limit", cfg_total_tags_limit)
        print_config("Max Tags Count", cfg_max_tags_count)
        # --- Print Scheduling Config ---
        print_config("Scheduling Mode", scheduling_mode)
        print_config("Schedule Interval (mins)", schedule_interval_minutes)
        if scheduling_mode == 'custom_tomorrow':
            custom_times_display = ", ".join([t.strftime('%I:%M %p') for t in parsed_config_times]) if parsed_config_times else "None"
            print_config("Custom Schedule Times", custom_times_display)
        print_config("Min Schedule Ahead (mins)", min_schedule_ahead_minutes)
        # --- Print Recording Config ---
        print_config("Enable Debug Recording", f"{Fore.GREEN}{enable_debug_recording}{Style.RESET_ALL}" if enable_debug_recording else f"{Fore.YELLOW}{enable_debug_recording}{Style.RESET_ALL}")
        if enable_debug_recording:
            print_config("FFmpeg Path", ffmpeg_path_config)
            print_config("Recording Output Dir", DEBUG_RECORDING_FOLDER)
        # --- End Print Config ---


        print_section_header("Initializing WebDriver"); driver = configure_driver()
        if not driver: raise Exception("WebDriver initialization failed. Check logs.")

        print_section_header("Loading Excel Data")
        excel_headers = {
            "Downloaded": ["Video ID", "Optimized Title", "Downloaded Date", "Views", "Uploader", "Original Title"],
            "Uploaded": ["Video Index", "Optimized Title", "YouTube Video ID", "Upload Timestamp", "Scheduled Time", "Publish Status"]
        }

        # Try to import excel_utils module
        try:
            import excel_utils
            excel_utils_available = True
            print_info("Using excel_utils module for robust Excel handling", indent=1)
        except ImportError:
            excel_utils_available = False
            print_warning("excel_utils module not available. Using fallback Excel handling.", indent=1)

        if excel_utils_available:
            # Use the excel_utils module for robust Excel handling
            try:
                wb, sheets, save_needed = excel_utils.load_or_create_excel(EXCEL_FILE_PATH, excel_headers)

                if not wb:
                    print_fatal(f"Failed to load or create Excel file: {EXCEL_FILE_PATH}")
                    raise Exception(f"Failed to load or create Excel file: {EXCEL_FILE_PATH}")

                downloaded_sheet = sheets.get("Downloaded")
                uploaded_sheet = sheets.get("Uploaded")

                # Save if needed using robust save mechanism
                if save_needed:
                    excel_save_required = True
                    if not excel_utils.safe_save_workbook(wb, EXCEL_FILE_PATH, close_excel=True, create_backup=True):
                        print_warning(f"Could not save structural changes to Excel. Will try again later.", indent=1)

                print_success("Excel loaded successfully using excel_utils.", indent=1)
            except Exception as e:
                print_error(f"Error using excel_utils: {e}", indent=1, include_traceback=True)
                print_warning("Falling back to standard Excel handling.", indent=1)
                excel_utils_available = False  # Force fallback

        # Fallback to original implementation if excel_utils is not available or failed
        if not excel_utils_available:
            if not os.path.exists(EXCEL_FILE_PATH):
                 print_info(f"Excel file '{EXCEL_FILE_PATH}' not found. Creating new workbook...", indent=1)
                 wb = Workbook()
                 downloaded_sheet = wb.active
                 downloaded_sheet.title = "Downloaded"
                 downloaded_sheet.append(excel_headers["Downloaded"])
                 uploaded_sheet = wb.create_sheet("Uploaded")
                 uploaded_sheet.append(excel_headers["Uploaded"])
                 excel_save_required = True

                 # Try to save with a simple retry mechanism
                 max_retries = 3
                 for attempt in range(max_retries):
                     try:
                         wb.save(EXCEL_FILE_PATH)
                         print_success(f"Created and saved new Excel file: {EXCEL_FILE_PATH} (attempt {attempt+1})", indent=1)
                         break
                     except PermissionError as pe:
                         if attempt < max_retries - 1:
                             print_warning(f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}", indent=1)
                             print_info(f"Retrying in 2 seconds...", indent=1)
                             time.sleep(2)
                         else:
                             print_fatal(f"Failed to save newly created Excel file '{EXCEL_FILE_PATH}' after {max_retries} attempts: {pe}")
                             raise
                     except Exception as e:
                         print_fatal(f"Failed to save newly created Excel file '{EXCEL_FILE_PATH}': {e}")
                         raise
            else:
                 try:
                     print_info(f"Loading existing Excel file: {EXCEL_FILE_PATH}", indent=1)
                     wb = load_workbook(EXCEL_FILE_PATH)
                     if "Downloaded" not in wb.sheetnames:
                         print_warning("'Downloaded' sheet missing. Creating...", indent=2)
                         downloaded_sheet = wb.create_sheet("Downloaded", 0)
                         downloaded_sheet.append(excel_headers["Downloaded"])
                         excel_save_required = True
                     else:
                         downloaded_sheet = wb["Downloaded"]

                     if "Uploaded" not in wb.sheetnames:
                         print_warning("'Uploaded' sheet missing. Creating...", indent=2)
                         uploaded_sheet = wb.create_sheet("Uploaded")
                         uploaded_sheet.append(excel_headers["Uploaded"])
                         excel_save_required = True
                     else:
                         uploaded_sheet = wb["Uploaded"]

                     print_success("Excel file loaded successfully.", indent=1)
                 except Exception as e:
                     print_fatal(f"Error loading existing Excel file '{EXCEL_FILE_PATH}': {e}")
                     raise

        print_section_header("Checking Scheduled Video Status")
        if check_and_update_scheduled(uploaded_sheet): excel_save_required = True

        uploaded_count = 0
        # --- Scheduling Tracking Variables ---
        last_schedule_time: Optional[datetime] = None # Tracks the last successfully calculated schedule time for interval calc
        first_video_this_run = True # Flag for Mode A's publish now logic
        # Make a mutable copy of parsed times for this run's 'custom_tomorrow' mode
        remaining_config_times: List[dt_time] = list(parsed_config_times) if scheduling_mode == 'custom_tomorrow' else []
        if scheduling_mode == 'custom_tomorrow':
            print_info(f"Starting run with {len(remaining_config_times)} custom time slots available.", indent=1)
        # --- End Scheduling Tracking Variables ---

        print_section_header("Scanning for Videos to Upload"); print_info(f"Maximum uploads for this run set to: {max_uploads}", indent=1); all_metadata_files = []
        try:
            if not os.path.isdir(METADATA_FOLDER): print_warning(f"Metadata folder '{METADATA_FOLDER}' not found.", indent=1)
            else:
                metadata_files_raw = [f for f in os.listdir(METADATA_FOLDER) if f.lower().endswith('.json') and re.match(r'video\d+\.json', f, re.IGNORECASE)]
                def get_video_index_from_filename(filename): match = re.search(r'video(\d+)\.json', filename, re.IGNORECASE); return int(match.group(1)) if match else float('inf')
                all_metadata_files = sorted(metadata_files_raw, key=get_video_index_from_filename)
                if all_metadata_files: print_success(f"Found {len(all_metadata_files)} potential metadata files to process.", indent=1)
                else: print_info("No metadata files matching 'video*.json' found.", indent=1)
        except Exception as e: print_fatal(f"Error scanning metadata folder '{METADATA_FOLDER}': {e}"); raise

        if all_metadata_files: print_section_header(f"Starting Upload Loop (Max: {max_uploads})")
        else: print_info("No videos to upload based on scan results.")

        # --- Main Upload Loop ---
        for metadata_file in all_metadata_files:
            if uploaded_count >= max_uploads: print_info(f"Reached maximum upload limit ({max_uploads}). Stopping."); break

            # --- Video File & Metadata Loading ---
            video_index_match = re.search(r'video(\d+)\.json', metadata_file, re.IGNORECASE)
            if not video_index_match: print_warning(f"Skipping file with unexpected name format: {metadata_file}", indent=1); continue

            video_index = video_index_match.group(1); metadata_path = os.path.join(METADATA_FOLDER, metadata_file); video_file_name = f"video{video_index}.mp4"; video_file_path = os.path.join(UPLOAD_FOLDER, video_file_name)
            print_info(f"--- Processing Video Index: {video_index} ({metadata_file}) ---", indent=1)

            if not os.path.exists(video_file_path): print_error(f"Video file not found: '{video_file_path}'. Skipping.", indent=2); continue

            metadata: Dict = {}
            try:
                with open(metadata_path, "r", encoding="utf-8") as f: metadata = json.load(f)
                metadata['video_index'] = video_index; # Ensure index is in metadata for upload function logging
                print_success("Metadata loaded successfully.", indent=2)
                if 'optimized_title' not in metadata or not metadata['optimized_title']: print_warning("Metadata missing 'optimized_title' or it's empty.", indent=3)
            except json.JSONDecodeError as je: print_error(f"Invalid JSON in {metadata_file}: {je}. Skipping.", indent=2); log_error_to_file(f"ERROR: JSON Decode Error in {metadata_file}: {je}", include_traceback=False); continue
            except Exception as e: print_error(f"Error reading {metadata_file}: {e}. Skipping.", indent=2); log_error_to_file(f"ERROR: Reading metadata {metadata_file}: {e}", include_traceback=True); continue
            # --- End Video File & Metadata Loading ---

            # --- Scheduling Logic ---
            publish_this_video_now = False # Default assumption
            target_schedule_time: Optional[datetime] = None # The final schedule time for this video
            action_desc = "" # Description for logging
            now = datetime.now()
            min_future_time = now + timedelta(minutes=min_schedule_ahead_minutes)
            tomorrow_date = now.date() + timedelta(days=1) # Calculate tomorrow's date

            print_info(f"Determining schedule for video {video_index} using mode: '{scheduling_mode}'", indent=2)

            if scheduling_mode == 'default_interval':
                # --- Mode A Logic ---
                if first_video_this_run:
                    publish_this_video_now = True
                    action_desc = "Publish Now (Mode A - First Video)"
                    # last_schedule_time remains None for the first published video
                else:
                    publish_this_video_now = False
                    # Calculate schedule time based on interval
                    if last_schedule_time is None:
                        # First video *to be scheduled* in Mode A. Base interval off 'now'.
                        calculated_time = now + timedelta(minutes=schedule_interval_minutes)
                    else:
                        # Schedule relative to the previous scheduled video
                        calculated_time = last_schedule_time + timedelta(minutes=schedule_interval_minutes)

                    # Ensure calculated time is valid (sufficiently in the future)
                    target_schedule_time = max(calculated_time, min_future_time)
                    action_desc = f"Schedule Interval (Mode A) for {target_schedule_time:%Y-%m-%d %H:%M:%S}"
                    last_schedule_time = target_schedule_time # Update for the next interval calculation

            elif scheduling_mode == 'custom_tomorrow':
                # --- Mode B Logic (Custom Time -> Interval Fallback) ---
                publish_this_video_now = False # Never publish immediately
                schedule_reason = ""

                # 1. Try Custom Config Time Slot (ALWAYS for TOMORROW onwards)
                if remaining_config_times:
                    current_config_time_slot = remaining_config_times[0]
                    print_info(f"Attempting to use config time slot: {current_config_time_slot.strftime('%I:%M %p')} starting from tomorrow ({tomorrow_date:%Y-%m-%d})", indent=3)

                    potential_schedule_tomorrow = datetime.combine(tomorrow_date, current_config_time_slot)
                    is_valid = False
                    if potential_schedule_tomorrow >= min_future_time and (last_schedule_time is None or potential_schedule_tomorrow > last_schedule_time):
                        target_schedule_time = potential_schedule_tomorrow
                        is_valid = True
                        print_info(f"Valid schedule found for tomorrow using config slot: {target_schedule_time:%Y-%m-%d %H:%M:%S}", indent=4)

                    if is_valid:
                        schedule_reason = f"Config Time Slot for Tomorrow ({target_schedule_time:%I:%M %p on %Y-%m-%d})"
                        remaining_config_times.pop(0) # Consume this time slot
                        print_info(f"Consumed time slot. {len(remaining_config_times)} slots remaining.", indent=4)
                    else:
                         print_warning(f"Config time slot {current_config_time_slot.strftime('%I:%M %p')} for tomorrow ({potential_schedule_tomorrow:%Y-%m-%d %H:%M:%S}) is invalid (too soon or conflicts with last schedule). Trying interval fallback.", indent=4)
                         # DO NOT consume the time slot if it wasn't valid

                # 2. Fallback to Interval Scheduling
                if target_schedule_time is None:
                    print_info("Trying 'Interval Fallback' rule.", indent=3)
                    if last_schedule_time is None:
                        # First video in Mode B to use interval (or custom slots failed). Base interval off 'now'.
                        calculated_time = now + timedelta(minutes=schedule_interval_minutes)
                    else:
                        # Schedule relative to the last valid schedule time (custom or previous interval)
                        calculated_time = last_schedule_time + timedelta(minutes=schedule_interval_minutes)

                    # Ensure calculated time is valid
                    target_schedule_time = max(calculated_time, min_future_time)
                    schedule_reason = f"Interval Fallback ({target_schedule_time:%Y-%m-%d %H:%M:%S})"
                    print_info(f"Valid schedule found using Interval rule: {target_schedule_time:%Y-%m-%d %H:%M:%S}", indent=4)

                # Final assignment and update last_schedule_time for Mode B
                if target_schedule_time:
                    last_schedule_time = target_schedule_time # Update for next interval calc
                    action_desc = f"Schedule (Mode B - Reason: {schedule_reason})"
                else:
                    # This case should ideally not happen if interval fallback works, but handle defensively
                    print_error(f"FATAL: Could not determine any valid schedule time for video {video_index} in Mode B. Skipping upload.", indent=2)
                    continue # Skip this video

            # --- End Scheduling Logic ---

            print_info(f"Action determined: {action_desc}", indent=2)
            if not publish_this_video_now and target_schedule_time:
                 print_info(f"Target Schedule Time: {target_schedule_time:%Y-%m-%d %H:%M:%S}", indent=2)
            elif publish_this_video_now:
                 print_info("Target Action: Publish Immediately", indent=2)
            else:
                 print_error("Error: No valid action (Publish or Schedule) determined. Skipping.", indent=2)
                 continue

            # --- >>> NEW UPLOAD ATTEMPT LOOP (Total 3 Attempts) <<< ---
            max_total_attempts = 3  # Total attempts: 1 initial + 2 retries
            final_upload_successful = False
            captured_youtube_video_id = None
            attempted_this_video_metric = False  # Track if attempt metric was incremented

            for current_attempt in range(1, max_total_attempts + 1):  # Loop 1, 2, 3
                print_info(f"\n--- Upload Attempt {current_attempt}/{max_total_attempts} for Video Index: {video_index} ---", indent=1)
                if current_attempt > 1:
                    print_info(f"Waiting before retry attempt {current_attempt}...", indent=2)
                    time.sleep(random.uniform(5, 10))  # Wait before retry with some randomness

                # --- Recording Logic ---
                local_recording_process: Optional[subprocess.Popen] = None
                local_recording_filename: Optional[str] = None
                error_during_this_attempt = False  # Track errors specifically in this attempt

                if enable_debug_recording:
                    start_result = start_recording(video_index, ffmpeg_path_config, driver)
                    if start_result:
                        local_recording_process, local_recording_filename = start_result
                    else:
                        print_warning(f"Could not start recording for attempt {current_attempt}.", indent=2)
                # --- End Recording Start ---

                try:
                    # Increment attempt metric only ONCE per video index, on the FIRST attempt
                    if current_attempt == 1 and not attempted_this_video_metric:
                        metrics["total_uploads_attempted"] += 1
                        attempted_this_video_metric = True

                    # --- Call upload_video ---
                    # Reset ID for this attempt
                    attempt_youtube_video_id = upload_video(
                        driver,
                        video_file_path,
                        metadata,
                        publish_now=publish_this_video_now,
                        schedule_time=target_schedule_time,
                        desc_limit=cfg_desc_limit,
                        tag_char_limit=cfg_tag_limit,
                        total_char_limit=cfg_total_tags_limit,
                        max_count_limit=cfg_max_tags_count
                    )

                    # --- Check Success for THIS attempt ---
                    if attempt_youtube_video_id:
                        print_success(f"Upload Attempt {current_attempt} SUCCEEDED (YT ID: {attempt_youtube_video_id}).", indent=2)
                        final_upload_successful = True
                        captured_youtube_video_id = attempt_youtube_video_id
                        # BREAK the attempt loop on success
                        break
                    else:
                        print_error(f"Upload Attempt {current_attempt} FAILED (No YT ID returned).", indent=2)
                        error_during_this_attempt = True  # Mark error for recording

                except (NoSuchWindowException, InvalidSessionIdException) as critical_wd_error:
                    error_during_this_attempt = True
                    final_upload_successful = False  # Ensure failure state
                    print_error(f"CRITICAL BROWSER SESSION ERROR during attempt {current_attempt}: {critical_wd_error}", indent=1)
                    log_error_to_file(f"CRITICAL WebDriver error during upload attempt {current_attempt} for {video_index}: {critical_wd_error}", include_traceback=True)
                    # For critical errors, break the attempt loop immediately and re-raise to trigger main finally block
                    if local_recording_process:
                        stop_recording(local_recording_process, local_recording_filename, keep_file=True)  # Stop recording before raising
                    raise critical_wd_error

                except Exception as upload_err:
                    error_during_this_attempt = True
                    final_upload_successful = False  # Ensure failure state
                    print_error(f"Exception during upload attempt {current_attempt}: {upload_err}", indent=1, include_traceback=True)
                    log_error_to_file(f"Exception during upload attempt {current_attempt} for {video_index}: {upload_err}", include_traceback=True)

                finally:
                    # --- Stop Recording Logic ---
                    if local_recording_process:
                        # Keep recording if attempt failed OR overall upload not yet successful
                        keep_the_recording = error_during_this_attempt or not final_upload_successful
                        stop_recording(local_recording_process, local_recording_filename, keep_the_recording)
                    # --- End Stop Recording ---

                # If this attempt failed but we have more attempts, continue to the next one
                if not final_upload_successful and current_attempt < max_total_attempts:
                    print_info(f"Will try again. {max_total_attempts - current_attempt} attempts remaining.", indent=2)

            # --- End of UPLOAD ATTEMPT LOOP ---

            # If all attempts failed, log the final failure
            if not final_upload_successful:
                print_error(f"All {max_total_attempts} upload attempts FAILED for video index {video_index}.", indent=1)
                log_error_to_file(f"ERROR: All {max_total_attempts} upload attempts failed for video index {video_index}.", step="retry_handler")

            # --- Post-Upload Actions ---
            # Check for both successful upload AND valid YouTube ID
            if final_upload_successful and captured_youtube_video_id:
                print_success(f"Upload validated successfully for video index {video_index} (YT ID: {captured_youtube_video_id})", indent=1)
                uploaded_count += 1
                metrics["total_uploads_successful"] += 1
                excel_save_required = True
                status = "Published" if publish_this_video_now else "Scheduled"
                # Pass target_schedule_time to Excel update function only if it was scheduled
                actual_schedule_time_for_excel = target_schedule_time if not publish_this_video_now else None

                # --- Add correlation data BEFORE deleting files ---
                discovery_keyword_for_cache = metadata.get("discovery_keyword")  # Get keyword before potential deletion
                add_to_correlation_cache(
                    video_index_str=f"video{video_index}",
                    discovery_keyword=discovery_keyword_for_cache,
                    youtube_video_id=captured_youtube_video_id  # Use the ID returned by upload_video
                )

                # Update Excel data
                update_excel_data(
                    downloaded_sheet, uploaded_sheet, video_index,
                    metadata.get('optimized_title', f'Video {video_index}'),
                    datetime.now(),  # Upload timestamp is always now
                    actual_schedule_time_for_excel,  # Pass the schedule time used (or None)
                    status,
                    youtube_video_id=captured_youtube_video_id  # Pass the captured YouTube Video ID
                )

                # Now delete the files (after correlation data has been saved)
                delete_uploaded_files(video_file_path, metadata_path)

                # Update last_schedule_time correctly
                if not publish_this_video_now and target_schedule_time:
                    last_schedule_time = target_schedule_time
                    print_info(f"Set last_schedule_time to: {last_schedule_time:%Y-%m-%d %H:%M:%S} for next calculation", indent=2)

                # Mark that the first video (if applicable in Mode A) is done
                if first_video_this_run:
                    first_video_this_run = False

                print_success(f"Successfully processed video index {video_index} (YT ID: {captured_youtube_video_id}). Run count: {uploaded_count}/{max_uploads}", indent=1)

            # --- ELSE: Handle the failure case explicitly ---
            else:
                # This runs if final_upload_successful is False OR captured_youtube_video_id is None/False
                print_error(f"Upload FAILED for video index {video_index}. Files NOT deleted. Check logs and YT Studio.", indent=1)
                log_error_to_file(f"Upload FAILED (final state) for video index {video_index}. Files kept.", step="post_upload_check", video_index=video_index)

                # Log the specific failure reason
                if not final_upload_successful:
                    print_error("Reason: Upload process did not complete successfully", indent=2)
                elif not captured_youtube_video_id:
                    print_error("Reason: No valid YouTube Video ID was captured", indent=2)

                # DO NOT update excel or delete files here
                print_warning("Continuing to the next video (if any).", indent=1)

            # Save metrics after each upload attempt
            save_performance_metrics(metrics)

            print_info(f"--- End Processing Video Index: {video_index} ---", indent=1); mimic_human_action_delay(2, 5) # Wait between uploads
        # --- End Main Upload Loop ---

        print_section_header("Finished Processing All Found Videos")
        if uploaded_count == 0 and all_metadata_files: print_info("No videos were successfully uploaded in this run.", indent=1)
        elif uploaded_count > 0: print_success(f"Successfully uploaded {uploaded_count} video(s) in this run.", indent=1)

    # --- Main Exception Handling ---
    except (WebDriverException, NoSuchWindowException, InvalidSessionIdException) as wd_e:
        print_fatal(f"Critical WebDriver error in main loop: {wd_e}", include_traceback=False)
        log_error_to_file(f"FATAL WebDriver error in main: {wd_e}", include_traceback=True)
    except KeyboardInterrupt:
        print_fatal("\nScript interrupted by user (Ctrl+C). Exiting.", log_to_file=False)
        log_error_to_file("Script interrupted by user (Ctrl+C).")
    except Exception as e:
        print_fatal(f"An unexpected error occurred in main loop: {e}", include_traceback=False)
        log_error_to_file(f"FATAL error in main loop: {e}", include_traceback=True)

    # --- Global Finally Block (ensures cleanup) ---
    finally:
        # --- Ensure any lingering recording is stopped ---
        if _current_recording_process:
            print_warning("Script exiting with an active recording process. Attempting emergency stop.", indent=1)
            # Always keep the recording file if the script exits unexpectedly
            stop_recording(_current_recording_process, _current_recording_filename, keep_file=True)
        # --- End Recording Cleanup ---

        if wb and excel_save_required:
            print_section_header("Saving Excel Data")

            # Try to import excel_utils module
            try:
                import excel_utils
                excel_utils_available = True
                print_info("Using excel_utils module for robust Excel saving", indent=1)
            except ImportError:
                excel_utils_available = False
                print_warning("excel_utils module not available. Using fallback Excel saving.", indent=1)

            if excel_utils_available:
                # Extract workbook data for backup in case save fails
                def extract_data(wb):
                    data = {}
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        data[sheet_name] = []
                        for row in sheet.iter_rows(values_only=True):
                            data[sheet_name].append(list(row))
                    return data

                # Use the robust save mechanism
                if excel_utils.save_workbook_with_fallback(wb, EXCEL_FILE_PATH, extract_data):
                    print_success(f"Excel data saved successfully to: {EXCEL_FILE_PATH} using excel_utils", indent=1)
                else:
                    # If all save methods failed, create a JSON backup
                    backup_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), f"excel_backup_data_{datetime.now():%Y%m%d_%H%M%S}.json")
                    print_warning(f"All Excel save methods failed. Creating JSON backup: {backup_file}", indent=1)
                    try:
                        with open(backup_file, "w", encoding='utf-8') as bf:
                            json.dump(extract_data(wb), bf, indent=4, default=str)
                            print_success(f"Saved backup to {backup_file}", indent=1)
                            log_error_to_file(f"Saved Excel data backup to JSON: {backup_file} after all Excel save methods failed.")
                    except Exception as be:
                        print_error(f"CRITICAL: Failed backup save: {be}", indent=1)
                        log_error_to_file(f"ERROR: Failed to save Excel backup: {be}", include_traceback=True)
            else:
                # Fallback to original save mechanism with simple retry
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        wb.save(EXCEL_FILE_PATH)
                        print_success(f"Excel data saved successfully to: {EXCEL_FILE_PATH} (attempt {attempt+1})", indent=1)
                        break
                    except PermissionError as pe:
                        if attempt < max_retries - 1:
                            msg = f"PermissionError saving Excel (attempt {attempt+1}/{max_retries}): {pe}"
                            print_warning(msg, indent=1)
                            log_error_to_file(f"WARNING: {msg}")
                            print_info(f"Retrying in 2 seconds...", indent=1)
                            time.sleep(2)
                        else:
                            # On last attempt failure, try to save to a backup file
                            msg = f"Failed to save Excel after {max_retries} attempts: {pe}"
                            print_error(msg, indent=1)
                            log_error_to_file(f"ERROR: {msg}")
                            backup_path = EXCEL_FILE_PATH + f".backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                            try:
                                wb.save(backup_path)
                                print_warning(f"Saved backup Excel: {backup_path}", indent=1)
                                log_error_to_file(f"Warning: Saved backup Excel to {backup_path} after primary save failed.")
                            except Exception as backup_e:
                                print_error(f"Failed to save backup Excel file: {backup_e}", indent=1)
                                log_error_to_file(f"ERROR: Failed to save backup Excel: {backup_e}")
                    except Exception as e:
                        msg = f"Error saving Excel file '{EXCEL_FILE_PATH}': {e}"
                        print_error(msg, indent=1)
                        log_error_to_file(f"ERROR: {msg}", include_traceback=True)
                        # Try to save to a backup file
                        backup_path = EXCEL_FILE_PATH + f".backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                        try:
                            wb.save(backup_path)
                            print_warning(f"Saved backup Excel: {backup_path}", indent=1)
                            log_error_to_file(f"Warning: Saved backup Excel to {backup_path} after primary save failed.")
                        except Exception as backup_e:
                            print_error(f"Failed to save backup Excel file: {backup_e}", indent=1)
                            log_error_to_file(f"ERROR: Failed to save backup Excel: {backup_e}")
                        break
        elif wb:
            print_section_header("Excel Data")
            print_info("No changes detected in Excel data requiring save.", indent=1)

        if driver:
            print_section_header("Shutting Down WebDriver")
            try: driver.quit(); print_success("WebDriver closed successfully.", indent=1)
            except (WebDriverException, InvalidSessionIdException) as qe: print_warning(f"Error quitting WebDriver (session might be invalid): {qe}", indent=1); log_error_to_file(f"Warning: Error quitting WebDriver: {qe}")
            except Exception as e: print_error(f"Unexpected error quitting WebDriver: {e}", indent=1); log_error_to_file(f"ERROR: Unexpected error quitting WebDriver: {e}", include_traceback=True)

        end_time = time.time(); duration = timedelta(seconds=end_time - start_time)
        print_section_header("Script Execution Finished"); print_info(f"Total execution time: {str(duration).split('.')[0]}", indent=1)
        if os.path.exists(ERROR_LOG_FILE) and os.path.getsize(ERROR_LOG_FILE) > 0: print_warning(f"Errors or warnings logged. Check log file: {ERROR_LOG_FILE}", indent=1)
        else: print_success("Script completed without logging errors.", indent=1)

        # Check if we should run an analysis based on error rate
        if genai is not None:
            # Load metrics to check error rate
            try:
                metrics = load_performance_metrics()
                if metrics["total_uploads_attempted"] > 0:
                    error_rate = (metrics["total_uploads_attempted"] - metrics["total_uploads_successful"]) / metrics["total_uploads_attempted"]
                    if error_rate >= MIN_ERROR_RATE_FOR_ANALYSIS and metrics["total_errors"] >= MIN_ERRORS_FOR_ANALYSIS:
                        print_info(f"Error rate is {error_rate:.1%}, which is above the threshold of {MIN_ERROR_RATE_FOR_ANALYSIS:.1%}.")
                        print_info(f"Consider running with --analyze flag to get AI-powered suggestions for improvement.")
                        print_info(f"Command: python \"uploader editing.py\" --analyze")
            except Exception as e:
                print_warning(f"Could not check error metrics: {e}")

# --- Script Entry Point ---
if __name__ == "__main__":
    try:
        # Check if we're in analyze mode
        analyze_mode = len(sys.argv) > 1 and sys.argv[1].lower() in ["--analyze", "-a"]

        # Only rotate/archive the log if it's getting too large and we're not in analyze mode
        if os.path.exists(ERROR_LOG_FILE) and not analyze_mode:
            try:
                file_size_mb = os.path.getsize(ERROR_LOG_FILE) / (1024 * 1024)
                # If file is larger than 5MB, archive it instead of clearing
                if file_size_mb > 5:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    archive_path = f"{ERROR_LOG_FILE}.{timestamp}.bak"
                    # Copy to archive file
                    with open(ERROR_LOG_FILE, "r", encoding="utf-8") as src, open(archive_path, "w", encoding="utf-8") as dst:
                        dst.write(src.read())
                    # Add header to current log but don't clear it completely
                    with open(ERROR_LOG_FILE, "w", encoding="utf-8") as f:
                        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Error log rotated. Previous logs archived to {archive_path}\n")
                    print_info(f"Error log rotated. Previous logs archived to {archive_path}")
            except Exception as e:
                print(f"Warning: Error managing log file '{ERROR_LOG_FILE}': {e}")
    except Exception as e:
        print(f"Warning: Error checking log file: {e}")

    main()
# --- End of the script ---